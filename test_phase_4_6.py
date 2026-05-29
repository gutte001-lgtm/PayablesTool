"""
test_phase_4_6.py -- Phase 4.6 reporting-alignment tests.

Plain-script style (mirrors test_phase_4_5.py): check(label, cond); exit code ==
failures; run with `python test_phase_4_6.py`.

PURE (sqlite/dicts only): the three /summary sections contain the right cells
(Right Now AP excludes debt_service + not_real_ap; debt_service is never in any
AP total), pipeline bucketing, the BOD report excludes debt/not-real, the
tightened deposit heuristic, and sync update-path debt detection guarded by
classified_by IS NULL.

ROUTE (Flask client): /classifications access model + mark-reviewed; the export
fence drops a not_due line from a check-run export. Live payables.db is never
opened. Route tests need SECRET_KEY in .env.
"""
import io
import shutil
import sqlite3
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

from dotenv import dotenv_values

FAILURES = []
_TMP = []


def check(label, cond):
    print(("ok   " if cond else "FAIL ") + label)
    if not cond:
        FAILURES.append(label)


ROOT = Path(__file__).resolve().parent
import summary      # noqa: E402
import triage       # noqa: E402
import sync         # noqa: E402
import init_db      # noqa: E402
import db           # noqa: E402
import openpyxl     # noqa: E402

TODAY = date(2026, 5, 28)


def mkrow(bid, obl, due_state, open_cents, *, vendor="V", invoice=None,
          expected=None, cat=None):
    return {"qb_bill_id": bid, "vendor": vendor + bid, "open_balance_cents": open_cents,
            "obligation_type": obl, "due_state": due_state, "invoice_due_date": invoice,
            "expected_payment_date": expected, "app_category": cat,
            "bill_date": "2026-04-01", "due_date": invoice}


# ====================================================================
print("=" * 60); print("test_summary_sections"); print("=" * 60)
ROWS = [
    mkrow("RN", "ordinary_ap", "due", 1000, invoice="2026-05-01", cat="Freight"),
    mkrow("WAIT", "ordinary_ap", "not_due", 2000, expected="2026-06-02"),
    mkrow("NR", "not_real_ap", "not_due", 3000),
    mkrow("DEBT", "debt_service", "not_due", 100000, invoice="2026-06-15"),
    mkrow("DEBTDUE", "debt_service", "due", 50000, invoice="2026-05-20"),
]
rn = summary.compute_right_now_ap(ROWS, TODAY)
pipe = summary.compute_pipeline(ROWS, TODAY)
debt = summary.compute_debt_service(ROWS)

check("rnap: total = ordinary_ap+due only ($1000)", rn["total_cents"] == 1000)
check("rnap: excludes debt_service and not_real_ap", rn["bill_count"] == 1)
check("pipeline: = ordinary_ap+not_due ($2000) + not_real_ap ($3000) = $5000",
      pipe["total_cents"] == 5000)
check("pipeline: waiting sub-total ($2000) and not_real sub-total ($3000) split",
      pipe["waiting_cents"] == 2000 and pipe["not_real_cents"] == 3000)
check("debt: total = both debt bills ($150000), regardless of due_state",
      debt["total_cents"] == 150000)
check("debt: due/not_due split ($50000 / $100000)",
      debt["due_cents"] == 50000 and debt["not_due_cents"] == 100000)
# the asymmetry: debt_service is NEVER in any AP total
check("asymmetry: debt_service $ is in neither Right Now AP nor Pipeline",
      1000 + 5000 == summary._open_total([r for r in ROWS if r["obligation_type"] != "debt_service"]))

sec = summary.split_sections(ROWS)
check("split: every row lands in exactly one section",
      len(sec["right_now"]) + len(sec["pipeline"]) + len(sec["debt_service"]) == len(ROWS))
check("split: no row in two sections",
      not (set(b["qb_bill_id"] for b in sec["right_now"])
           & set(b["qb_bill_id"] for b in sec["pipeline"])))

# ---- pipeline bucket boundaries ----
print("=" * 60); print("test_pipeline_buckets"); print("=" * 60)
def pb(days):
    return summary.pipeline_bucket((TODAY + timedelta(days=days)).isoformat(), TODAY)
check("bucket: past date folds into This week", pb(-5) == "This week")
check("bucket: +7d -> This week", pb(7) == "This week")
check("bucket: +8d -> This month", pb(8) == "This month")
check("bucket: +30d -> This month", pb(30) == "This month")
check("bucket: +31d -> Next month", pb(31) == "Next month")
check("bucket: +60d -> Next month", pb(60) == "Next month")
check("bucket: +61d -> Later", pb(61) == "Later")
check("bucket: NULL -> No date set", summary.pipeline_bucket(None, TODAY) == "No date set")

# ---- BOD report excludes debt + not-real ----
print("=" * 60); print("test_bod_report_excludes_debt_and_notreal"); print("=" * 60)
wb = summary.build_bod_ap_workbook(ROWS, TODAY, "2026-05-28 12:00")
bio = io.BytesIO(); wb.save(bio); bio.seek(0)
rwb = openpyxl.load_workbook(bio)
ws = rwb["BOD AP Report"]
cells = [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 4)]
ordinary_total_row = [ws.cell(r, 3).value for r in range(1, ws.max_row + 1)
                      if (ws.cell(r, 1).value or "").startswith("ORDINARY AP TOTAL")]
check("bod: ordinary-AP total = $30.00 (1000+2000 cents), excludes debt+not_real",
      ordinary_total_row == [30.0])
check("bod: debt vendor (VDEBT) absent from the report",
      not any("VDEBT" in str(v) for v in cells))
check("bod: not_real vendor (VNR) absent from the report",
      not any("VNR" in str(v) for v in cells))

# ---- tightened deposit heuristic ----
print("=" * 60); print("test_deposit_heuristic_tightened"); print("=" * 60)
candela = {"app_category": "New Device Purchases", "has_credit_applied": 0}
b_partial = {"amount_cents": 10000, "open_balance_cents": 4000}  # partially paid
check("heuristic: Candela-style partially-paid COGS -> ordinary_ap (NOT deposit)",
      triage.suggest(b_partial, [], candela)[0] == "ordinary_ap")
check("heuristic: ...and flags 'possible deposit?' for a human look",
      "possible deposit" in triage.suggest(b_partial, [], candela)[3])
with_credit = {"app_category": "New Device Purchases", "has_credit_applied": 1}
check("heuristic: vendor credit present -> not_real_ap/deposit",
      triage.suggest({"amount_cents": 10000, "open_balance_cents": 4000}, [], with_credit)[:3]
      == ("not_real_ap", "not_due", "deposit"))
check("heuristic: liability line still -> debt_service",
      triage.suggest({"amount_cents": 1, "open_balance_cents": 1},
                     [{"gl_account_number_canonical": "26110", "gl_account_number": "26110"}],
                     {"app_category": None, "has_credit_applied": 0})[0] == "debt_service")

# ---- sync update-path detection guarded by classified_by IS NULL ----
print("=" * 60); print("test_update_path_detection_guard"); print("=" * 60)


def fresh_sqlite():
    p = Path(tempfile.mkdtemp()); _TMP.append(p)
    cn = sqlite3.connect(p / "t.db"); cn.row_factory = sqlite3.Row
    cn.executescript(init_db.SCHEMA)
    cn.execute("INSERT INTO users (username,name,role,is_active) VALUES ('joe','Joe','controller',1)")
    cn.commit()
    return cn


def add_bill_meta(cn, bid, obligation="ordinary_ap", classified_by=None):
    cn.execute("INSERT INTO bill (qb_bill_id,vendor,amount_cents,open_balance_cents,"
               "due_date,last_synced_at) VALUES (?,?,?,?,?,?)",
               (bid, "V" + bid, 10000, 10000, "2026-06-15", "t"))
    cn.execute("INSERT INTO bill_metadata (qb_bill_id,approval_state,obligation_type,"
               "due_state,invoice_due_date,classified_by,classified_at,created_at,updated_at) "
               "VALUES (?,?,?,?,?,?,?,?,?)",
               (bid, "Controller_Reviewed", obligation, "not_due", "2026-06-15",
                classified_by, ("t" if classified_by else None), "t", "t"))


cn = fresh_sqlite()
# untouched bill (classified_by NULL) that now hits a liability account
add_bill_meta(cn, "AUTO", obligation="ordinary_ap", classified_by=None)
created, detected, _ = sync._ensure_metadata(cn, "AUTO", None, None, False, "now",
                                             invoice_due_date="2026-06-15", is_debt_service=True)
cn.commit()
row = cn.execute("SELECT obligation_type FROM bill_metadata WHERE qb_bill_id='AUTO'").fetchone()
check("update-path: untouched ordinary_ap + liability hit -> auto debt_service",
      row["obligation_type"] == "debt_service" and detected is True)
check("update-path: auto-detection wrote a classification_audit row",
      cn.execute("SELECT COUNT(*) FROM classification_audit WHERE bill_id='AUTO' "
                 "AND field='obligation_type' AND to_value='debt_service'").fetchone()[0] == 1)

# human-classified bill (classified_by set) hitting a liability account: LEFT ALONE
add_bill_meta(cn, "HUMAN", obligation="ordinary_ap", classified_by=1)
_c, detected2, _ = sync._ensure_metadata(cn, "HUMAN", None, None, False, "now",
                                        invoice_due_date="2026-06-15", is_debt_service=True)
cn.commit()
row = cn.execute("SELECT obligation_type FROM bill_metadata WHERE qb_bill_id='HUMAN'").fetchone()
check("update-path GUARD: human-classified bill is NEVER overwritten by sync",
      row["obligation_type"] == "ordinary_ap" and detected2 is False)
check("update-path GUARD: no classification_audit row written for the human bill",
      cn.execute("SELECT COUNT(*) FROM classification_audit WHERE bill_id='HUMAN'").fetchone()[0] == 0)
cn.close()


# ====================================================================
# ROUTE TESTS
# ====================================================================
if not dotenv_values(ROOT / ".env").get("SECRET_KEY"):
    print("\nSKIPPING ROUTE TESTS: no SECRET_KEY in .env")
else:
    from werkzeug.security import generate_password_hash  # noqa: E402
    PW = generate_password_hash("testpw")
    USERS = [("marilyn", "Marilyn", "ap_clerk"), ("joe", "Joe", "controller"),
             ("shaun", "Shaun", "cfo")]

    def fresh_route_db():
        d = Path(tempfile.mkdtemp()); _TMP.append(d)
        db.DB_PATH = d / "p46.db"
        cn = sqlite3.connect(db.DB_PATH); cn.executescript(init_db.SCHEMA)
        for u, name, role in USERS:
            cn.execute("INSERT INTO users (username,name,role,password_hash,is_active) "
                       "VALUES (?,?,?,?,1)", (u, name, role, PW))
        init_db.seed_classification_reasons(cn)  # classify route validates reasons
        cn.commit(); cn.close()

    from app import app  # noqa: E402
    app.config["WTF_CSRF_ENABLED"] = False
    app.config["TESTING"] = True

    def client(u):
        c = app.test_client(); c.post("/login", data={"username": u, "password": "testpw"}); return c

    def rconn():
        cn = sqlite3.connect(db.DB_PATH); cn.row_factory = sqlite3.Row; return cn

    print("=" * 60); print("test_classifications_access"); print("=" * 60)
    fresh_route_db()
    check("access: ap_clerk -> 403 on /classifications",
          client("marilyn").get("/classifications").status_code == 403)
    check("access: controller -> 200", client("joe").get("/classifications").status_code == 200)
    check("access: cfo -> 200", client("shaun").get("/classifications").status_code == 200)
    check("access: ap_clerk -> 403 on BOD report",
          client("marilyn").get("/summary/bod-ap-report.xlsx").status_code == 403)
    check("access: cfo -> 200 on BOD report",
          client("shaun").get("/summary/bod-ap-report.xlsx").status_code == 200)

    # mark-reviewed writes the per-user audit marker and resets the list
    shaun = client("shaun")
    shaun.post("/classifications/mark-reviewed")
    cn = rconn()
    n = cn.execute("SELECT COUNT(*) FROM audit_log WHERE action='classifications_reviewed' "
                   "AND user_id=(SELECT id FROM users WHERE username='shaun')").fetchone()[0]
    cn.close()
    check("mark-reviewed: writes a classifications_reviewed audit_log row", n == 1)

    print("=" * 60); print("test_export_fence_drops_not_due_line"); print("=" * 60)
    fresh_route_db()
    cn = rconn()

    def seed(bid, obl, due_state):
        cn.execute("INSERT INTO bill (qb_bill_id,vendor,bill_number,amount_cents,"
                   "open_balance_cents,bill_date,due_date,is_paid,last_synced_at) "
                   "VALUES (?,?,?,?,?,?,?,0,?)",
                   (bid, "Vendor" + bid, "B" + bid, 10000, 10000, "2026-05-01",
                    "2026-05-15", "t"))
        cn.execute("INSERT INTO bill_metadata (qb_bill_id,approval_state,obligation_type,"
                   "due_state,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                   (bid, "Controller_Reviewed", obl, due_state, "t", "t"))
    seed("ELIG", "ordinary_ap", "due")       # belongs on the export
    seed("NOTDUE", "ordinary_ap", "not_due")  # must be fenced OUT of the export
    rid = cn.execute("INSERT INTO pay_run (name,status,created_by,created_at,updated_at) "
                     "VALUES ('R','Locked',2,'t','t')").lastrowid
    for bid in ("ELIG", "NOTDUE"):
        cn.execute("INSERT INTO pay_run_line (pay_run_id,qb_bill_id,payment_method,"
                   "amount_to_pay_cents,included,line_state) VALUES (?,?,?,?,1,'Pending')",
                   (rid, bid, "Check", 10000))
    cn.commit(); cn.close()

    resp = client("joe").get(f"/pay-runs/{rid}/export/cfo.xlsx")
    check("export: cfo export 200 for the Locked run", resp.status_code == 200)
    ewb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ews = ewb.active
    vals = [ews.cell(r, 1).value for r in range(1, ews.max_row + 1)]
    check("export fence: the due bill (VendorELIG) IS on the export",
          any("VendorELIG" in str(v) for v in vals))
    check("export fence: the not_due bill (VendorNOTDUE) is NOT on the export",
          not any("VendorNOTDUE" in str(v) for v in vals))

    # ---------- item-6 guard: adversarial sequences (folded from scratch) ------
    # Each drives the REAL classify route (stamps classified_by) and a realistic
    # sync (bill_reduces_liability -> _ensure_metadata update path -> promote).
    print("=" * 60); print("test_item6_guard_adversarial"); print("=" * 60)
    import classifications  # noqa: E402
    FUTURE_INV, PAST_INV = "2099-01-01", "2020-01-01"
    TODAY_ISO = date.today().isoformat()

    def rc():
        cn = sqlite3.connect(db.DB_PATH); cn.row_factory = sqlite3.Row; return cn

    def adv_line(bid, acct):
        cn = rc()
        cn.execute("DELETE FROM bill_line WHERE qb_bill_id=?", (bid,))
        cn.execute("INSERT INTO bill_line (qb_bill_id,line_number,gl_account_name,"
                   "gl_account_number,gl_account_number_canonical,line_amount_cents) "
                   "VALUES (?,?,?,?,?,?)", (bid, 1, acct, acct, acct, 10000))
        cn.commit(); cn.close()

    def adv_seed(bid, acct, invoice=FUTURE_INV):
        cn = rc()
        cn.execute("INSERT INTO bill (qb_bill_id,vendor,bill_number,amount_cents,"
                   "open_balance_cents,bill_date,due_date,is_paid,last_synced_at) "
                   "VALUES (?,?,?,?,?,?,?,0,?)",
                   (bid, "V" + bid, "B" + bid, 10000, 10000, "2026-04-01", invoice, "t"))
        cn.execute("INSERT INTO bill_metadata (qb_bill_id,approval_state,obligation_type,"
                   "due_state,invoice_due_date,expected_payment_date,created_at,updated_at) "
                   "VALUES (?,?,?,?,?,?,?,?)",
                   (bid, "Controller_Reviewed", "ordinary_ap", "not_due", invoice,
                    invoice, "t", "t"))
        cn.commit(); cn.close()
        adv_line(bid, acct)

    def adv_sync(bid):
        cn = rc()
        lines = [dict(r) for r in cn.execute(
            "SELECT gl_account_number, gl_account_number_canonical FROM bill_line "
            "WHERE qb_bill_id=?", (bid,))]
        is_debt = sync.bill_reduces_liability(lines)
        idd = cn.execute("SELECT invoice_due_date FROM bill_metadata WHERE qb_bill_id=?",
                         (bid,)).fetchone()[0]
        sync._ensure_metadata(cn, bid, None, None, False, sync._now_iso(),
                              invoice_due_date=idd, is_debt_service=is_debt)
        sync.promote_debt_service_due(cn, TODAY_ISO, sync._now_iso())
        cn.commit(); cn.close()

    def adv_st(bid):
        cn = rc()
        m = cn.execute("SELECT obligation_type,due_state,classified_by FROM bill_metadata "
                       "WHERE qb_bill_id=?", (bid,)).fetchone()
        n = cn.execute("SELECT COUNT(*) FROM classification_audit WHERE bill_id=? "
                       "AND field='obligation_type'", (bid,)).fetchone()[0]
        cn.close()
        return m["obligation_type"], m["due_state"], m["classified_by"], n

    fresh_route_db()

    # 1 — human override survives repeated syncs
    adv_seed("A1", "26110"); adv_sync("A1")
    check("adv1: auto-detect -> debt_service", adv_st("A1")[0] == "debt_service")
    client("joe").post("/bills/A1/classify", data={"obligation_type": "ordinary_ap"})
    for _ in range(3):
        adv_sync("A1")
    o, _, cb, _ = adv_st("A1")
    check("adv1: human ordinary_ap survives 3 syncs", o == "ordinary_ap")
    check("adv1: human stamp persists", cb is not None)

    # 2 — auto-detect idempotent, one audit row
    adv_seed("A2", "26110"); adv_sync("A2")
    for _ in range(5):
        adv_sync("A2")
    o, _, _, n = adv_st("A2")
    check("adv2: stays debt_service", o == "debt_service")
    check("adv2: exactly ONE obligation_type audit row (no per-sync spam)", n == 1)

    # 3 — auto-detect leaves classified_by NULL; human choice protected
    adv_seed("A3", "26110"); adv_sync("A3")
    o, _, cb, _ = adv_st("A3")
    check("adv3: auto-detect leaves classified_by NULL", cb is None and o == "debt_service")
    client("joe").post("/bills/A3/classify", data={"obligation_type": "ordinary_ap"})
    for _ in range(2):
        adv_sync("A3")
    o, _, cb, _ = adv_st("A3")
    check("adv3: human choice protected through syncs", o == "ordinary_ap" and cb is not None)

    # 4a — human-classified bill not flipped by recode to 26110
    adv_seed("A4", "56100")  # non-liability
    client("joe").post("/bills/A4/classify",
                       data={"obligation_type": "ordinary_ap", "due_state": "due",
                             "classification_reason": "other"})
    adv_line("A4", "26110")  # vendor recode -> now hits liability
    adv_sync("A4")
    o, _, cb, _ = adv_st("A4")
    check("adv4a: human-classified bill NOT flipped by recode", o == "ordinary_ap" and cb is not None)

    # 5a — clerk due_state flip on debt bill; obligation untouched
    adv_seed("A5", "26110"); adv_sync("A5")
    client("marilyn").post("/bills/A5/classify",
                           data={"obligation_type": "debt_service", "due_state": "due",
                                 "classification_reason": "debt_service"})
    adv_sync("A5")
    o, d, cb, _ = adv_st("A5")
    check("adv5a: obligation stays debt_service", o == "debt_service")
    check("adv5a: due_state stays the human's value (due)", d == "due")
    check("adv5a: clerk stamp persists", cb is not None)

    # 5b — INTENDED zero-tolerance override + Cleanup C (stamp survives)
    adv_seed("A6", "26110", invoice=PAST_INV); adv_sync("A6")  # past -> promote -> due
    client("marilyn").post("/bills/A6/classify",
                           data={"obligation_type": "debt_service", "due_state": "not_due",
                                 "classification_reason": "debt_service"})  # human "hold"
    o, d, cb, _ = adv_st("A6")
    check("adv5b: human set due_state=not_due and is stamped", d == "not_due" and cb is not None)
    human_id = cb
    adv_sync("A6")  # promote overrides (zero-tolerance)
    o, d, cb, _ = adv_st("A6")
    check("adv5b INTENDED: past-due debt not_due overridden to due by promote",
          d == "due" and o == "debt_service")
    check("adv5b CLEANUP C: classified_by SURVIVES the override (not nulled)",
          cb == human_id)
    cn = rc()
    overridden = {r["bill_id"] for r in classifications._system_overrides(cn)}
    cn.close()
    check("adv5b VISIBILITY B: the override surfaces on /classifications", "A6" in overridden)


# ====================================================================
for d in _TMP:
    shutil.rmtree(d, ignore_errors=True)
print()
if FAILURES:
    print(f"{len(FAILURES)} FAILURE(S): " + "; ".join(FAILURES))
else:
    print("ALL PHASE 4.6 CHECKS PASSED")
sys.exit(len(FAILURES))
