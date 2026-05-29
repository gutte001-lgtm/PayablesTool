"""
test_ceo_workpaper.py -- the three-tier CEO payment workpaper.

Plain-script style (mirrors test_picker_diagnostic.py / test_phase_5_export.py):
check(label, cond); exit code == failures; run `python test_ceo_workpaper.py`.

PART A (pure data): payruns.held_and_notdue_tiers() tier assignment on a fixture
  that exercises every path -- eligible-parked, deselected, rejected-this-run,
  rejected-on-another-run, claimed-elsewhere, paid-this-run, not-yet-due, still-
  in-processing (New/AP), not_real_ap, excluded-classification, closed.

PART B (pure render): excel_payrun.group_by_category + build_ceo_workpaper_workbook
  -- category subtotals tie to source rows; the workbook reloads cleanly.
PART C (routes): temp DB + Flask client -- access gating (controller+cfo vs 403),
  pre-lock (CFO_Approved) allowed + Draft refused, audit row, paid-tier grand
  total == the run's payable total. Skipped if no SECRET_KEY in .env.

The live payables.db and the real exports/ dir are never touched.
"""
import io
import json
import shutil
import sqlite3
import sys
import tempfile
from pathlib import Path

FAILURES = []
_TMP = []


def check(label, cond):
    print(("ok   " if cond else "FAIL ") + label)
    if not cond:
        FAILURES.append(label)


ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
import init_db   # noqa: E402
import bills      # noqa: E402
import payruns    # noqa: E402
import excel_payrun  # noqa: E402
import openpyxl   # noqa: E402


def _new_path():
    d = Path(tempfile.mkdtemp()); _TMP.append(d)
    return d / "t.db"


def _conn(p):
    cn = sqlite3.connect(p); cn.row_factory = sqlite3.Row
    return cn


def add_bill(cn, bid, approval="Controller_Reviewed", due="due",
             obligation="ordinary_ap", classification=None, open_bal=10000,
             cat="New Device Purchases", ok_for_ceo=0):
    cn.execute("INSERT INTO bill (qb_bill_id,vendor,bill_number,amount_cents,"
               "open_balance_cents,bill_date,due_date,is_paid,last_synced_at) "
               "VALUES (?,?,?,?,?,?,?,?,?)",
               (bid, "V" + bid, "B" + bid, 10000, open_bal, "2026-05-01",
                "2026-06-15", 0 if open_bal > 0 else 1, "t"))
    cn.execute("INSERT INTO bill_metadata (qb_bill_id,app_category,approval_state,"
               "classification,obligation_type,due_state,ok_for_ceo,created_at,"
               "updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
               (bid, cat, approval, classification, obligation, due, ok_for_ceo,
                "t", "t"))


def add_line(cn, run_id, bid, included=1, line_state="Pending", method="Check",
             amount=10000, cfo_note=None):
    cn.execute("INSERT INTO pay_run_line (pay_run_id,qb_bill_id,payment_method,"
               "amount_to_pay_cents,included,line_state,cfo_note) "
               "VALUES (?,?,?,?,?,?,?)",
               (run_id, bid, method, amount, included, line_state, cfo_note))


def fixture():
    """run 1 = the workpaper's run; run 2 = another run."""
    p = _new_path(); cn = _conn(p)
    cn.executescript(init_db.SCHEMA)
    for rid in (1, 2):
        cn.execute("INSERT INTO pay_run (id,name,status,created_at,updated_at) "
                   "VALUES (?,?,?,?,?)", (rid, "R%d" % rid, "Draft", "t", "t"))
    # eligible, not on any run -> HELD A (parked)
    add_bill(cn, "PARKED", cat="Occupancy")
    # eligible, claimed by run 2 (included, non-rejected) -> NOT held
    add_bill(cn, "CLAIMED_OTHER"); add_line(cn, 2, "CLAIMED_OTHER")
    # eligible, payable on THIS run (claimed) -> tier 1, NOT held
    add_bill(cn, "PAID_THIS"); add_line(cn, 1, "PAID_THIS")
    # eligible, deselected on THIS run (included=0 releases it) -> HELD A
    add_bill(cn, "DESELECTED", cat="Legal Fees"); add_line(cn, 1, "DESELECTED", included=0)
    # eligible, rejected on THIS run -> HELD B (reason = cfo_note)
    add_bill(cn, "REJ_THIS"); add_line(cn, 1, "REJ_THIS", line_state="Rejected",
                                       cfo_note="vendor dispute")
    # eligible, rejected on ANOTHER run -> released -> HELD A
    add_bill(cn, "REJ_OTHER"); add_line(cn, 2, "REJ_OTHER", line_state="Rejected",
                                        cfo_note="other-run reject")
    # reviewed real AP but not yet due -> NOT YET DUE
    add_bill(cn, "NOTDUE", due="not_due", cat="Parts & Products")
    # still in processing -> footnote count only
    add_bill(cn, "NEWBILL", approval="New")
    add_bill(cn, "APREV", approval="AP_Reviewed")
    # never appears anywhere
    add_bill(cn, "NOTREAL", obligation="not_real_ap")
    add_bill(cn, "EXCLCLASS", classification="Refund-Visibility")
    add_bill(cn, "CLOSED", open_bal=0)
    cn.commit()
    return p, cn


print("=" * 60); print("PART A -- tier assignment (pure data)"); print("=" * 60)
p, cn = fixture()
t = payruns.held_and_notdue_tiers(cn, 1)
held_ids = sorted(r["qb_bill_id"] for r in t["held"])
notdue_ids = sorted(r["qb_bill_id"] for r in t["not_yet_due"])

check("held == {PARKED, DESELECTED, REJ_OTHER, REJ_THIS}",
      held_ids == ["DESELECTED", "PARKED", "REJ_OTHER", "REJ_THIS"])
check("paid-on-this-run bill excluded from held", "PAID_THIS" not in held_ids)
check("claimed-by-another-run bill excluded from held", "CLAIMED_OTHER" not in held_ids)
check("rejected-this-run appears exactly once (no A/B double-count)",
      held_ids.count("REJ_THIS") == 1)

by_id = {r["qb_bill_id"]: r for r in t["held"]}
check("REJ_THIS is held_kind=rejected_this_run with its cfo_note reason",
      by_id["REJ_THIS"]["held_kind"] == "rejected_this_run"
      and by_id["REJ_THIS"]["reason"] == "vendor dispute")
check("PARKED is held_kind=eligible_parked with empty reason",
      by_id["PARKED"]["held_kind"] == "eligible_parked" and by_id["PARKED"]["reason"] == "")
check("DESELECTED (included=0) surfaces as eligible_parked",
      by_id["DESELECTED"]["held_kind"] == "eligible_parked")
check("REJ_OTHER (rejected on another run) surfaces as eligible_parked",
      by_id["REJ_OTHER"]["held_kind"] == "eligible_parked")

check("not_yet_due == {NOTDUE}", notdue_ids == ["NOTDUE"])
check("processing_count == 2 (NEWBILL + APREV)", t["processing_count"] == 2)

absent = set(held_ids) | set(notdue_ids)
check("not_real_ap never appears", "NOTREAL" not in absent)
check("excluded-classification never appears", "EXCLCLASS" not in absent)
check("closed (open_balance=0) never appears", "CLOSED" not in absent)
check("New/AP bills are not in held or not_yet_due (footnote only)",
      "NEWBILL" not in absent and "APREV" not in absent)
cn.close()


print("=" * 60); print("PART B -- category grouping + workbook (pure render)"); print("=" * 60)
p, cn = fixture()
t = payruns.held_and_notdue_tiers(cn, 1)
import exports as _exp  # noqa: E402
paid = excel_payrun.payable_rows(_exp._run_lines(cn, 1), ceo=False)  # just PAID_THIS
cn.close()

held_blocks = excel_payrun.group_by_category(t["held"], "open_balance_cents")
held_grand = next(b for b in held_blocks if b["kind"] == "grand_total")["subtotal_cents"]
check("held grand total == sum of source open balances",
      held_grand == sum(r["open_balance_cents"] for r in t["held"]))
check("held category subtotals sum to the grand total",
      sum(b["subtotal_cents"] for b in held_blocks if b["kind"] == "section") == held_grand)
cats = {b["label"] for b in held_blocks if b["kind"] == "section"}
check("held groups by app_category (Occupancy + Legal Fees + New Device Purchases)",
      {"Occupancy", "Legal Fees", "New Device Purchases"} <= cats)

wb = excel_payrun.build_ceo_workpaper_workbook(
    {"id": 1, "name": "R1", "week_ending": "2026-05-21"},
    paid, t["held"], t["not_yet_due"], "Joe", t["processing_count"])
bio = io.BytesIO(); wb.save(bio); bio.seek(0)
ws = openpyxl.load_workbook(bio)[excel_payrun.WORKPAPER_SHEET]
colA = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
check("workbook reloads cleanly + has the three tier bands",
      "PAID THIS RUN" in colA and "HELD BY CHOICE" in colA and "NOT YET DUE" in colA)
check("workbook footnote names the processing count",
      any("processing" in str(v).lower() for v in colA if v))


print("\n" + "=" * 60); print("PART C -- routes (temp DB + client)"); print("=" * 60)
from dotenv import dotenv_values  # noqa: E402
if not dotenv_values(ROOT / ".env").get("SECRET_KEY"):
    print("SKIP: no SECRET_KEY in .env (pure tests above still ran)")
else:
    from werkzeug.security import generate_password_hash  # noqa: E402
    import db          # noqa: E402
    import exports     # noqa: E402
    from app import app  # noqa: E402
    app.config["WTF_CSRF_ENABLED"] = False
    app.config["TESTING"] = True
    PW = generate_password_hash("testpw")
    USERS = [("marilyn", "Marilyn", "ap_clerk"), ("joe", "Joe", "controller"),
             ("shaun", "Shaun", "cfo")]
    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def fresh_db():
        d = Path(tempfile.mkdtemp()); _TMP.append(d)
        db.DB_PATH = d / "wp.db"
        c = sqlite3.connect(db.DB_PATH); c.executescript(init_db.SCHEMA)
        for u, name, role in USERS:
            c.execute("INSERT INTO users (username,name,role,password_hash,is_active) "
                      "VALUES (?,?,?,?,1)", (u, name, role, PW))
        c.commit(); c.close()
        exp = d / "exports"; exp.mkdir(); exports.EXPORTS_DIR = exp
        return exp

    def _cnr():
        c = sqlite3.connect(db.DB_PATH); c.row_factory = sqlite3.Row; return c

    def login(u):
        c = app.test_client(); c.post("/login", data={"username": u, "password": "testpw"}); return c

    def seed_run(status):
        c = _cnr()
        cur = c.execute("INSERT INTO pay_run (name,week_ending,created_by,status,"
                        "created_at,updated_at) VALUES (?,?,?,?,?,?)",
                        ("Run", "2026-05-22", None, status, "2026-05-22", "2026-05-22"))
        c.commit(); rid = cur.lastrowid; c.close(); return rid

    URL = "/pay-runs/{}/export/ceo-workpaper.xlsx"

    # ---- pre-lock (CFO_Approved) allowed; paid total ties to run payable total ----
    exp = fresh_db()
    c = _cnr()
    add_bill(c, "p1", cat="Occupancy"); add_bill(c, "p2", cat="Legal Fees")
    add_bill(c, "h1", cat="Parts & Products")          # eligible parked -> held
    c.commit(); c.close()
    rid = seed_run("CFO_Approved")
    c = _cnr(); add_line(c, rid, "p1", amount=10000); add_line(c, rid, "p2", amount=20000)
    c.commit(); c.close()
    r = login("shaun").get(URL.format(rid))
    check("route: CFO_Approved (pre-lock) export 200", r.status_code == 200)
    check("route: content-type is xlsx", r.headers.get("Content-Type", "").startswith(XLSX))
    f = list(exp.glob(f"PayRun_{rid}_CEO_Workpaper_2026-05-22_v01.xlsx"))
    check("route: versioned _v01 workpaper file written", len(f) == 1)
    ws = openpyxl.load_workbook(io.BytesIO(r.data))[excel_payrun.WORKPAPER_SHEET]
    colA = [ws.cell(rr, 1).value for rr in range(1, ws.max_row + 1)]
    paid_total, in_paid = None, False
    for rr in range(1, ws.max_row + 1):
        v = ws.cell(rr, 1).value
        if v == "PAID THIS RUN":
            in_paid = True
        elif v in ("HELD BY CHOICE", "NOT YET DUE"):
            in_paid = False
        elif in_paid and v == "TOTAL":
            paid_total = ws.cell(rr, excel_payrun._WP_AMOUNT_COL).value
    check("route: paid-tier TOTAL == run payable total (300.00)", paid_total == 300.0)
    check("route: held tier shows the parked bill's category (Parts & Products)",
          "Parts & Products" in colA)

    c = _cnr()
    a = c.execute("SELECT after FROM audit_log WHERE action='pay_run_exported' "
                  "ORDER BY id DESC LIMIT 1").fetchone()
    c.close()
    ap = json.loads(a["after"])
    check("route: audit row written with export=ceo_workpaper + paid_total_cents",
          ap.get("export") == "ceo_workpaper" and ap.get("generated_by") == "Shaun"
          and ap.get("paid_total_cents") == 30000)

    # ---- access gating ----
    check("route: controller allowed (200)", login("joe").get(URL.format(rid)).status_code == 200)
    check("route: ap_clerk forbidden (403)", login("marilyn").get(URL.format(rid)).status_code == 403)

    # ---- Draft refused, no file ----
    exp2 = fresh_db()
    c = _cnr(); add_bill(c, "d1", cat="Occupancy"); c.commit(); c.close()
    rid2 = seed_run("Draft")
    c = _cnr(); add_line(c, rid2, "d1"); c.commit(); c.close()
    rd = login("shaun").get(URL.format(rid2))
    check("route: Draft run refused (302), writes no file",
          rd.status_code == 302 and not list(exp2.glob("*.xlsx")))

# ====================================================================
for d in _TMP:
    shutil.rmtree(d, ignore_errors=True)
print()
if FAILURES:
    print(f"{len(FAILURES)} FAILURE(S): " + "; ".join(FAILURES))
else:
    print("ALL CEO-WORKPAPER CHECKS PASSED")
sys.exit(len(FAILURES))
