"""
excel_payrun.py -- Phase 5 PURE builders for the CFO Pay-Run Excel and the CEO
printout. No Flask, no DB: takes already-fetched pay_run_line row dicts and
returns an openpyxl Workbook, so the grouping + formatting are unit-testable in
isolation.

Section grouping (locked; first match wins, per bill):
  1. app_category startswith "Contractor - " -> by payment_method:
       Contractor Checks / Contractor Wire / Contractor Credit Cards / Contractor ACH
  2. app_category == "Pre-owned Device Purchases" -> "Buys"            (any method)
  3. app_category == "Refunds"                    -> "Refunds/Reimbursements" (any method)
  4. else by payment_method: Checks / Credit Cards / ACH-Wire
A blank/unknown payment_method falls into the Check bucket (Contractor Checks /
Checks) -- the default disbursement method; flagged in the build notes.

Subtotals sum amount_to_pay_cents ("Open balance"), NOT amount_cents ("Amount").
All written values are STATIC (no formulas) and dates are real Excel dates -- the
two bugs in the legacy workbook this rebuild fixes.
"""
from datetime import date, datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import bills  # reuse METHODS (do not redefine)

CONTRACTOR_PREFIX = "Contractor - "
SHEET_TITLE = "Pay Run"

HEADERS = ["Vendor", "Vendor Type", "Bill number", "Date", "Due date", "Amount",
           "Open balance", "Payment Method", "Bill Approval", "Approval Date",
           "Receipt/Delivery", "Memo", "Notes"]

# 1-based column indices
_MONEY_COLS = (6, 7)          # F Amount, G Open balance
_DATE_COLS = (4, 5, 10, 11)   # D Date, E Due date, J Approval Date, K Receipt/Delivery
_OPEN_BAL_COL = 7             # subtotals live in the Open balance column
_MEMO_COL = 12

_MONEY_FMT = "#,##0.00"
_ACCT_FMT = r'_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
_DATE_FMT = "mm/dd/yyyy"
_WIDTHS = {1: 30, 2: 25, 3: 14, 4: 11, 5: 11, 6: 12, 7: 13, 8: 14,
           9: 24, 10: 13, 11: 14, 12: 72, 13: 30}

# (bucket key, display label, subtotal label) -- order matters.
_CONTRACTOR_SECTIONS = [
    ("contractor_check", "Contractor Checks", "Contractor Checks Total"),
    ("contractor_wire", "Contractor Wire", "Contractor Wire"),
    ("contractor_cc", "Contractor Credit Cards", "Contractor Credit Cards"),
    ("contractor_ach", "Contractor ACH", "Contractor ACH"),
]
_OTHER_SECTIONS = [
    ("checks", "Checks", "Checks"),
    ("buys", "Buys", "Buys"),
    ("refunds", "Refunds/Reimbursements", "Refunds/Reimbursements"),
    ("credit_cards", "Credit Cards", "Credit Cards"),
    ("ach_wire", "ACH/Wire", "ACH/Wire"),
]

_CONTRACTOR_BY_METHOD = {"Check": "contractor_check", "Wire": "contractor_wire",
                         "Credit Card": "contractor_cc", "ACH": "contractor_ach"}
_OTHER_BY_METHOD = {"Check": "checks", "Credit Card": "credit_cards",
                    "Wire": "ach_wire", "ACH": "ach_wire"}


# ----------------------------------------------------------------------
# Filtering + grouping (pure)
# ----------------------------------------------------------------------

def _is_payable(r):
    """Payable = included AND not Rejected (line_state in Pending/Approved)."""
    return bool(r.get("included")) and (r.get("line_state") != "Rejected")


def payable_rows(lines, ceo=False):
    """The rows that land in an export: payable, and (CEO) ok_for_ceo=1."""
    out = [r for r in lines if _is_payable(r)]
    if ceo:
        out = [r for r in out if r.get("ok_for_ceo") in (1, True)]
    return out


def _section_key(row):
    cat = row.get("app_category") or ""
    method = row.get("payment_method")
    if cat.startswith(CONTRACTOR_PREFIX):
        return _CONTRACTOR_BY_METHOD.get(method, "contractor_check")
    if cat == "Pre-owned Device Purchases":
        return "buys"
    if cat == "Refunds":
        return "refunds"
    return _OTHER_BY_METHOD.get(method, "checks")


def group_into_sections(lines):
    """Group payable rows into the locked 11-section structure.

    Returns an ordered list of blocks (already-payable rows assumed):
      {"kind": "section"|"contractor_total"|"grand_total",
       "label": <display label or None>, "subtotal_label": <str>,
       "rows": [row dicts], "subtotal_cents": <int>}
    Empty sections are omitted. "Contractor Total" appears only if any contractor
    section has rows. "TOTAL" sums the detail-section subtotals (it does NOT
    re-add Contractor Total, so contractor $ isn't double-counted).
    """
    buckets = {}
    for r in lines:
        buckets.setdefault(_section_key(r), []).append(r)

    def _sort(rows):
        return sorted(rows, key=lambda r: ((r.get("app_category") or ""),
                                           (r.get("vendor") or ""),
                                           (r.get("bill_number") or "")))

    def _sub(rows):
        return sum((r.get("amount_to_pay_cents") or 0) for r in rows)

    blocks, contractor_total, any_contractor = [], 0, False
    for key, label, sublabel in _CONTRACTOR_SECTIONS:
        rows = buckets.get(key)
        if rows:
            any_contractor = True
            rows = _sort(rows); st = _sub(rows); contractor_total += st
            blocks.append({"kind": "section", "label": label,
                           "subtotal_label": sublabel, "rows": rows,
                           "subtotal_cents": st})
    if any_contractor:
        blocks.append({"kind": "contractor_total", "label": None,
                       "subtotal_label": "Contractor Total", "rows": [],
                       "subtotal_cents": contractor_total})
    for key, label, sublabel in _OTHER_SECTIONS:
        rows = buckets.get(key)
        if rows:
            rows = _sort(rows); st = _sub(rows)
            blocks.append({"kind": "section", "label": label,
                           "subtotal_label": sublabel, "rows": rows,
                           "subtotal_cents": st})
    grand = sum(b["subtotal_cents"] for b in blocks if b["kind"] == "section")
    blocks.append({"kind": "grand_total", "label": None, "subtotal_label": "TOTAL",
                   "rows": [], "subtotal_cents": grand})
    return blocks


# ----------------------------------------------------------------------
# Cell value helpers (pure)
# ----------------------------------------------------------------------

def _xl_date(s):
    """ISO 'YYYY-MM-DD' (or date/datetime) -> datetime.date for a real Excel date;
    None / unparseable -> None (blank cell)."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    try:
        return date.fromisoformat(str(s)[:10])
    except ValueError:
        return None


def _approval(row):
    a = (row.get("approver_name") or "").strip()
    c = (row.get("approval_channel") or "").strip()
    if a and c:
        return f"{a} - {c}"
    return a or c or ""


def _row_values(row):
    return [
        row.get("vendor") or "",
        row.get("app_category") or "",
        row.get("bill_number") or "",
        _xl_date(row.get("bill_date")),
        _xl_date(row.get("due_date")),
        (row.get("amount_cents") or 0) / 100.0,
        (row.get("amount_to_pay_cents") or 0) / 100.0,
        row.get("payment_method") or "",
        _approval(row),
        _xl_date(row.get("approval_date")),
        _xl_date(row.get("receipt_delivery_date")),
        row.get("qb_memo") or "",
        row.get("cfo_note") or "",
    ]


def _now_iso():
    return datetime.now(timezone.utc).isoformat(sep=" ", timespec="seconds")


# ----------------------------------------------------------------------
# Workbook builders
# ----------------------------------------------------------------------

def _build(run, payable, generated_by, ceo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    sub_font = Font(name="Arial", size=11, bold=True)
    foot_font = Font(name="Arial", size=8, italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(HEADERS, start=1):
        ws.cell(1, c, h).font = header_font
    ws.freeze_panes = "A2"
    for c, w in _WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    def _write_subtotal(r, label, cents):
        lc = ws.cell(r, 1, label); lc.font = sub_font
        vc = ws.cell(r, _OPEN_BAL_COL, cents / 100.0)  # STATIC value, not a formula
        vc.font = sub_font; vc.number_format = _ACCT_FMT

    r = 2
    for b in group_into_sections(payable):
        if b["kind"] == "section":
            for row in b["rows"]:
                for c, v in enumerate(_row_values(row), start=1):
                    cell = ws.cell(r, c, v); cell.font = body_font
                    if c in _MONEY_COLS:
                        cell.number_format = _MONEY_FMT
                    elif c in _DATE_COLS and v is not None:
                        cell.number_format = _DATE_FMT
                    elif c == _MEMO_COL:
                        cell.alignment = wrap
                r += 1
        _write_subtotal(r, b["subtotal_label"], b["subtotal_cents"])
        r += 1

    # in-sheet footer stamp
    r += 1
    ws.cell(r, 1, f"PayRun #{run['id']} · {run['name']} · exported "
                  f"{_now_iso()} by {generated_by}").font = foot_font

    if ceo:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_title_rows = "1:1"
        ws.oddFooter.center.text = "Page &P of &N"
        ws.oddFooter.left.text = f"PayRun #{run['id']} exported {date.today().isoformat()}"
    return wb


def build_cfo_workbook(run, lines, generated_by):
    """CFO Pay-Run Excel: all payable lines, grouped + subtotaled."""
    return _build(run, payable_rows(lines, ceo=False), generated_by, ceo=False)


def build_ceo_workbook(run, lines, generated_by):
    """CEO printout: payable lines filtered to ok_for_ceo=1, landscape print setup."""
    return _build(run, payable_rows(lines, ceo=True), generated_by, ceo=True)


# ----------------------------------------------------------------------
# CEO payment workpaper (three tiers, grouped by app_category)
# ----------------------------------------------------------------------
#
# A separate single-sheet workbook from the CFO/CEO check-run exports above.
# All three tiers share ONE category grouping (not the method-based
# group_into_sections, which keys off payment_method that held/not-yet-due bills
# -- not on a run -- don't have). Amounts: PAID THIS RUN subtotals on
# amount_to_pay_cents (so the tier total ties to the run's payable total); HELD
# and NOT YET DUE subtotal on open_balance_cents (what is being held / coming due).
# Same openpyxl conventions as _build: static values, real Excel dates, the
# shared _MONEY_FMT / _ACCT_FMT / _DATE_FMT.

WORKPAPER_SHEET = "CEO Workpaper"
_WP_HEADERS = ["Vendor", "Category", "Bill number", "Date", "Due date", "Amount",
               "CEO visible", "Note / reason"]
_WP_AMOUNT_COL = 6
_WP_DATE_COLS = (4, 5)
_WP_WIDTHS = {1: 30, 2: 26, 3: 14, 4: 11, 5: 11, 6: 14, 7: 11, 8: 48}
_UNCATEGORIZED = "Uncategorized"


def group_by_category(rows, amount_field):
    """Group rows by app_category (blank -> 'Uncategorized'), summing
    `amount_field`. Returns ordered blocks like group_into_sections:
      {"kind": "section"|"grand_total", "label"/"subtotal_label": <category|TOTAL>,
       "rows": [...], "subtotal_cents": <int>}
    Categories sorted alphabetically; rows within by vendor, bill_number."""
    buckets = {}
    for r in rows:
        buckets.setdefault((r.get("app_category") or _UNCATEGORIZED), []).append(r)

    blocks = []
    for cat in sorted(buckets):
        cat_rows = sorted(buckets[cat], key=lambda r: ((r.get("vendor") or ""),
                                                       (r.get("bill_number") or "")))
        st = sum((r.get(amount_field) or 0) for r in cat_rows)
        blocks.append({"kind": "section", "label": cat, "subtotal_label": cat,
                       "rows": cat_rows, "subtotal_cents": st})
    grand = sum(b["subtotal_cents"] for b in blocks)
    blocks.append({"kind": "grand_total", "label": None, "subtotal_label": "TOTAL",
                   "rows": [], "subtotal_cents": grand})
    return blocks


def _wp_row_values(row, amount_field, show_ceo, show_reason):
    return [
        row.get("vendor") or "",
        row.get("app_category") or _UNCATEGORIZED,
        row.get("bill_number") or "",
        _xl_date(row.get("bill_date")),
        _xl_date(row.get("due_date")),
        (row.get(amount_field) or 0) / 100.0,
        ("Yes" if row.get("ok_for_ceo") in (1, True) else "No") if show_ceo else "",
        (row.get("reason") or "") if show_reason else "",
    ]


def build_ceo_workpaper_workbook(run, paid_rows, held_rows, notdue_rows,
                                 generated_by, processing_count):
    """Single-sheet CEO workpaper: PAID THIS RUN / HELD BY CHOICE / NOT YET DUE,
    each grouped by category with subtotals + a tier TOTAL, then the
    in-processing footnote count. Landscape print setup like the CEO printout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKPAPER_SHEET
    header_font = Font(name="Arial", size=10, bold=True)
    band_font = Font(name="Arial", size=12, bold=True)
    body_font = Font(name="Arial", size=10)
    sub_font = Font(name="Arial", size=11, bold=True)
    foot_font = Font(name="Arial", size=8, italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(_WP_HEADERS, start=1):
        ws.cell(1, c, h).font = header_font
    ws.freeze_panes = "A2"
    for c, w in _WP_WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    def _amount(r, cents, font):
        cell = ws.cell(r, _WP_AMOUNT_COL, cents / 100.0)
        cell.font = font; cell.number_format = _ACCT_FMT

    # (band label, rows, amount_field, show_ceo_flag, show_reason)
    tiers = [
        ("PAID THIS RUN", paid_rows, "amount_to_pay_cents", True, False),
        ("HELD BY CHOICE", held_rows, "open_balance_cents", False, True),
        ("NOT YET DUE", notdue_rows, "open_balance_cents", False, False),
    ]

    r = 2
    for label, rows, amount_field, show_ceo, show_reason in tiers:
        ws.cell(r, 1, label).font = band_font
        r += 1
        for b in group_by_category(rows, amount_field):
            if b["kind"] == "section":
                for row in b["rows"]:
                    for c, v in enumerate(
                            _wp_row_values(row, amount_field, show_ceo, show_reason), start=1):
                        cell = ws.cell(r, c, v); cell.font = body_font
                        if c == _WP_AMOUNT_COL:
                            cell.number_format = _MONEY_FMT
                        elif c in _WP_DATE_COLS and v is not None:
                            cell.number_format = _DATE_FMT
                        elif c == 8:
                            cell.alignment = wrap
                    r += 1
                ws.cell(r, 1, b["subtotal_label"]).font = sub_font
                _amount(r, b["subtotal_cents"], sub_font)
                r += 1
            else:  # grand_total
                ws.cell(r, 1, b["subtotal_label"]).font = sub_font
                _amount(r, b["subtotal_cents"], sub_font)
                r += 1
        r += 1  # blank row between tiers

    ws.cell(r, 1, f"{processing_count} bill(s) still in processing "
                  f"(New / AP-Reviewed), not shown.").font = foot_font
    r += 2
    ws.cell(r, 1, f"CEO Workpaper · PayRun #{run['id']} · {run['name']} · "
                  f"generated {_now_iso()} by {generated_by}").font = foot_font

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"
    ws.oddFooter.center.text = "Page &P of &N"
    return wb
