"""
test_phase_5_export.py -- Phase 5 export tests (CFO Pay-Run Excel + CEO printout).

Plain-script style (mirrors test_phase_4.py): check(label, cond); exit code ==
failures; run with `python test_phase_5_export.py`.

PART A (pure): build fixture line dicts, exercise excel_payrun grouping +
build_*_workbook, reload via openpyxl, assert section order / subtotals / static
values / typed dates / filters.
PART B (routes): temp DB + Flask test client -- versioned filenames, audit row,
role gating, not-Locked + empty-CEO redirects. Skipped if no SECRET_KEY in .env.
The live payables.db and the real exports/ dir are never touched.
"""
import io
import json
import shutil
import sqlite3
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

FAILURES = []
_TMP = []


def check(label, cond):
    print(("ok   " if cond else "FAIL ") + label)
    if not cond:
        FAILURES.append(label)


ROOT = Path(__file__).resolve().parent
import excel_payrun  # noqa: E402
import openpyxl      # noqa: E402

RUN = {"id": 7, "name": "Test Run", "week_ending": "2026-05-21"}


def mk(bid, cat, method, amount=10000, included=1, line_state="Pending",
       ok_for_ceo=0, cfo_note="", vendor=None):
    return {
        "qb_bill_id": bid, "app_category": cat, "payment_method": method,
        "amount_to_pay_cents": amount, "amount_cents": amount, "included": included,
        "line_state": line_state, "ok_for_ceo": ok_for_ceo, "cfo_note": cfo_note,
        "vendor": vendor or ("Vendor " + bid), "bill_number": "B-" + bid,
        "bill_date": "2026-05-01", "due_date": "2026-05-15",
        "approver_name": "Marilyn", "approval_channel": "Pur Board",
        "approval_date": "2026-04-28", "receipt_delivery_date": "2026-05-02",
        "qb_memo": "OPS-1 memo " + bid,
    }


LINES = [
    mk("c1", "Contractor - Service & Repair", "Check", 10000),
    mk("c2", "Contractor - Training", "Wire", 20000),
    mk("c3", "Contractor - Outside Sales Commissions", "Credit Card", 30000),
    mk("c4", "Contractor - Service & Repair", "ACH", 40000),
    mk("k1", "New Device Purchases", "Check", 50000),
    mk("b1", "Pre-owned Device Purchases", "Wire", 60000),     # category beats method
    mk("r1", "Refunds", "Check", 70000),
    mk("cc1", "Parts & Products", "Credit Card", 80000),
    mk("aw1", "Occupancy", "Wire", 90000),
    mk("aw2", "Legal Fees", "ACH", 5000),
    mk("rej", "New Device Purchases", "Check", 99999, line_state="Rejected"),
    mk("exc", "New Device Purchases", "Check", 88888, included=0),
]

EXPECTED = [  # (subtotal_label, dollars)
    ("Contractor Checks Total", 100.0), ("Contractor Wire", 200.0),
    ("Contractor Credit Cards", 300.0), ("Contractor ACH", 400.0),
    ("Contractor Total", 1000.0), ("Checks", 500.0), ("Buys", 600.0),
    ("Refunds/Reimbursements", 700.0), ("Credit Cards", 800.0),
    ("ACH/Wire", 950.0), ("TOTAL", 4550.0),
]
SUBTOTAL_LABELS = {lbl for lbl, _ in EXPECTED}

print("=" * 60); print("PART A -- pure grouping + workbook"); print("=" * 60)

blocks = excel_payrun.group_into_sections(excel_payrun.payable_rows(LINES))
got_seq = [(b["subtotal_label"], b["subtotal_cents"] / 100.0) for b in blocks]
check("group: section order + subtotals match expected", got_seq == EXPECTED)
check("group: grand TOTAL == sum of detail-section subtotals",
      blocks[-1]["subtotal_cents"] == sum(b["subtotal_cents"] for b in blocks if b["kind"] == "section"))

# section membership
by_label = {b["label"]: [r["qb_bill_id"] for r in b["rows"]] for b in blocks if b["kind"] == "section"}
check("group: contractor+Check -> Contractor Checks (not Checks)",
      "c1" in by_label["Contractor Checks"] and "c1" not in by_label.get("Checks", []))
check("group: pre-owned+Wire -> Buys (not ACH/Wire)",
      "b1" in by_label["Buys"] and "b1" not in by_label.get("ACH/Wire", []))
check("group: Refunds -> Refunds/Reimbursements",
      "r1" in by_label["Refunds/Reimbursements"])
check("group: ACH/Wire combines Wire + ACH",
      set(by_label["ACH/Wire"]) == {"aw1", "aw2"})

# filters: rejected + excluded never appear
all_ids = {rid for ids in by_label.values() for rid in ids}
check("group: rejected line excluded", "rej" not in all_ids)
check("group: excluded(included=0) line excluded", "exc" not in all_ids)

# blank/unknown method routing
check("group: blank method (non-contractor) -> Checks",
      excel_payrun._section_key(mk("x", "Occupancy", None)) == "checks")
check("group: blank method (contractor) -> Contractor Checks bucket",
      excel_payrun._section_key(mk("x", "Contractor - Training", None)) == "contractor_check")


def reload_wb(wb):
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return openpyxl.load_workbook(bio)


wb = excel_payrun.build_cfo_workbook(RUN, LINES, "Joe")
rwb = reload_wb(wb)               # also proves the workbook opens cleanly
ws = rwb["Pay Run"]
check("wb: sheet titled 'Pay Run'", rwb.sheetnames == ["Pay Run"])
check("wb: header row matches the 13 columns",
      [ws.cell(1, c).value for c in range(1, 14)] == excel_payrun.HEADERS)

# walk col A for subtotal label rows, in order, with their col G value
seen = []
for r in range(2, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a in SUBTOTAL_LABELS:
        seen.append((a, ws.cell(r, 7).value, ws.cell(r, 7).data_type))
check("wb: subtotal label order + values match (reloaded)",
      [(s[0], s[1]) for s in seen] == EXPECTED)
check("wb: subtotal cells are STATIC numbers (data_type != 'f')",
      all(s[2] != "f" and isinstance(s[1], (int, float)) for s in seen))

# first data row (row 2) types
check("wb: Date col D is a real Excel date (not string)",
      isinstance(ws.cell(2, 4).value, (datetime, date)))
check("wb: Due date col E is a real Excel date",
      isinstance(ws.cell(2, 5).value, (datetime, date)))
check("wb: Amount col F numeric with money format",
      isinstance(ws.cell(2, 6).value, (int, float)) and ws.cell(2, 6).number_format == "#,##0.00")

# CEO workbook: only ok_for_ceo=1
ceo_lines = [dict(x) for x in LINES]
for x in ceo_lines:
    x["ok_for_ceo"] = 1 if x["qb_bill_id"] in ("c1", "k1") else 0
cwb = reload_wb(excel_payrun.build_ceo_workbook(RUN, ceo_lines, "Shaun"))
cws = cwb["Pay Run"]
ceo_vendors = {cws.cell(r, 1).value for r in range(2, cws.max_row + 1)}
check("ceo: includes ok_for_ceo=1 bills (c1, k1)",
      "Vendor c1" in ceo_vendors and "Vendor k1" in ceo_vendors)
check("ceo: excludes ok_for_ceo=0 bills (b1, r1 absent)",
      "Vendor b1" not in ceo_vendors and "Vendor r1" not in ceo_vendors)
check("ceo: landscape print orientation", cws.page_setup.orientation == "landscape")
check("ceo: header repeated on each page (print_title_rows -> $1:$1)",
      (cws.print_title_rows or "").replace("$", "") == "1:1")

# ====================================================================
print("\n" + "=" * 60); print("PART B -- routes (temp DB + client)"); print("=" * 60)
from dotenv import dotenv_values  # noqa: E402
if not dotenv_values(ROOT / ".env").get("SECRET_KEY"):
    print("SKIP: no SECRET_KEY in .env (pure tests above still ran)")
else:
    from werkzeug.security import generate_password_hash  # noqa: E402
    import db          # noqa: E402
    import init_db     # noqa: E402
    import exports     # noqa: E402
    from app import app  # noqa: E402
    app.config["WTF_CSRF_ENABLED"] = False
    app.config["TESTING"] = True
    PW = generate_password_hash("testpw")
    USERS = [("marilyn", "Marilyn", "ap_clerk"), ("joe", "Joe", "controller"),
             ("shaun", "Shaun", "cfo")]

    def fresh_db():
        d = Path(tempfile.mkdtemp()); _TMP.append(d)
        db.DB_PATH = d / "p5.db"
        cn = sqlite3.connect(db.DB_PATH); cn.executescript(init_db.SCHEMA)
        for u, name, role in USERS:
            cn.execute("INSERT INTO users (username,name,role,password_hash,is_active) "
                       "VALUES (?,?,?,?,1)", (u, name, role, PW))
        cn.commit(); cn.close()
        exp = d / "exports"; exp.mkdir(); exports.EXPORTS_DIR = exp
        return exp

    def _cn():
        c = sqlite3.connect(db.DB_PATH); c.row_factory = sqlite3.Row; return c

    def uid(u):
        c = _cn(); r = c.execute("SELECT id FROM users WHERE username=?", (u,)).fetchone(); c.close(); return r["id"]

    def login(u):
        c = app.test_client(); c.post("/login", data={"username": u, "password": "testpw"}); return c

    def seed_bill(bid, cat, ok_for_ceo=0, amount=10000):
        c = _cn()
        c.execute("INSERT INTO bill (qb_bill_id,vendor,bill_number,amount_cents,"
                  "open_balance_cents,bill_date,due_date,is_paid,last_synced_at) "
                  "VALUES (?,?,?,?,?,?,?,0,?)",
                  (bid, "Vendor " + bid, "B-" + bid, amount, amount,
                   "2026-05-01", "2026-05-15", "2026-05-22"))
        # Phase 4.6 added an export fence (obligation_type IN ordinary_ap/
        # debt_service AND due_state='due'); these export tests assume the lines
        # are payable, so seed due_state='due' (obligation_type defaults to
        # ordinary_ap).
        c.execute("INSERT INTO bill_metadata (qb_bill_id,app_category,approval_state,"
                  "approver_name,approval_channel,approval_date,receipt_delivery_date,"
                  "ok_for_ceo,due_state,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,'due',?,?)",
                  (bid, cat, "Controller_Reviewed", "Marilyn", "Pur Board",
                   "2026-04-28", "2026-05-02", ok_for_ceo, "2026-05-01", "2026-05-01"))
        c.commit(); c.close()

    def seed_run(status="Locked"):
        c = _cn()
        cur = c.execute("INSERT INTO pay_run (name,week_ending,created_by,status,"
                        "created_at,updated_at) VALUES (?,?,?,?,?,?)",
                        ("Run", "2026-05-22", uid("joe"), status, "2026-05-22", "2026-05-22"))
        c.commit(); rid = cur.lastrowid; c.close(); return rid

    def seed_line(rid, bid, method="Check", amount=10000, included=1, line_state="Pending"):
        c = _cn()
        c.execute("INSERT INTO pay_run_line (pay_run_id,qb_bill_id,payment_method,"
                  "amount_to_pay_cents,included,line_state) VALUES (?,?,?,?,?,?)",
                  (rid, bid, method, amount, included, line_state))
        c.commit(); c.close()

    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ---- versioning + audit ----
    exp = fresh_db()
    seed_bill("d1", "Contractor - Service & Repair", ok_for_ceo=1)
    seed_bill("d2", "New Device Purchases", ok_for_ceo=0)
    rid = seed_run("Locked")
    seed_line(rid, "d1", "Check", 10000)
    seed_line(rid, "d2", "Wire", 20000)
    cfo = login("shaun")
    resp1 = cfo.get(f"/pay-runs/{rid}/export/cfo.xlsx")
    check("route: cfo export 200", resp1.status_code == 200)
    check("route: cfo content-type is xlsx", resp1.headers.get("Content-Type", "").startswith(XLSX))
    f1 = list(exp.glob(f"PayRun_{rid}_2026-05-22_v01.xlsx"))
    check("route: first cfo file is _v01", len(f1) == 1)
    resp2 = cfo.get(f"/pay-runs/{rid}/export/cfo.xlsx")
    f2 = list(exp.glob(f"PayRun_{rid}_2026-05-22_v02.xlsx"))
    check("route: second cfo export -> _v02 (no overwrite)", resp2.status_code == 200 and len(f2) == 1)
    # downloaded bytes open cleanly + carry both lines' grand total
    dwb = openpyxl.load_workbook(io.BytesIO(resp1.data))["Pay Run"]
    tot = [dwb.cell(r, 7).value for r in range(2, dwb.max_row + 1) if dwb.cell(r, 1).value == "TOTAL"]
    check("route: downloaded workbook TOTAL = 300.00 (100+200)", tot == [300.0])
    # audit
    c = _cn()
    arows = c.execute("SELECT after FROM audit_log WHERE action='pay_run_exported' ORDER BY id").fetchall()
    c.close()
    check("route: 2 pay_run_exported audit rows", len(arows) == 2)
    a0 = json.loads(arows[0]["after"])
    check("route: audit payload has export/filename/version/row_count/total_cents/generated_by",
          a0.get("export") == "cfo" and a0.get("version") == 1 and a0.get("row_count") == 2
          and a0.get("total_cents") == 30000 and a0.get("filename", "").endswith("_v01.xlsx")
          and a0.get("generated_by") == "Shaun")

    # ---- CEO export ----
    respc = cfo.get(f"/pay-runs/{rid}/export/ceo.xlsx")
    fc = list(exp.glob(f"PayRun_{rid}_CEO_2026-05-22_v01.xlsx"))
    check("route: ceo export 200 + _CEO_ file", respc.status_code == 200 and len(fc) == 1)
    cdwb = openpyxl.load_workbook(io.BytesIO(respc.data))["Pay Run"]
    cvendors = {cdwb.cell(r, 1).value for r in range(2, cdwb.max_row + 1)}
    check("route: ceo file has only ok_for_ceo bill (d1, not d2)",
          "Vendor d1" in cvendors and "Vendor d2" not in cvendors)

    # ---- role gating ----
    check("route: ap_clerk forbidden on cfo export (403)",
          login("marilyn").get(f"/pay-runs/{rid}/export/cfo.xlsx").status_code == 403)
    check("route: controller allowed on cfo export (200)",
          login("joe").get(f"/pay-runs/{rid}/export/cfo.xlsx").status_code == 200)
    check("route: controller forbidden on ceo export (403)",
          login("joe").get(f"/pay-runs/{rid}/export/ceo.xlsx").status_code == 403)

    # ---- empty CEO (all ok_for_ceo=0) -> redirect ----
    exp2 = fresh_db()
    seed_bill("e1", "New Device Purchases", ok_for_ceo=0)
    rid2 = seed_run("Locked"); seed_line(rid2, "e1", "Check", 10000)
    rce = login("shaun").get(f"/pay-runs/{rid2}/export/ceo.xlsx")
    check("route: empty-CEO export redirects (302), writes no file",
          rce.status_code == 302 and not list(exp2.glob("*CEO*")))

    # ---- not-Locked -> redirect, no file ----
    exp3 = fresh_db()
    seed_bill("n1", "New Device Purchases")
    rid3 = seed_run("Draft"); seed_line(rid3, "n1", "Check", 10000)
    rnl = login("shaun").get(f"/pay-runs/{rid3}/export/cfo.xlsx")
    check("route: not-Locked cfo export redirects (302), writes no file",
          rnl.status_code == 302 and not list(exp3.glob("*.xlsx")))

# ====================================================================
for d in _TMP:
    shutil.rmtree(d, ignore_errors=True)
print()
if FAILURES:
    print(f"{len(FAILURES)} FAILURE(S): " + "; ".join(FAILURES))
else:
    print("ALL PHASE 5 EXPORT CHECKS PASSED")
sys.exit(len(FAILURES))
