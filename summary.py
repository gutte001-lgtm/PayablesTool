"""
summary.py -- Phase 6 Spend Summary dashboard at /summary.

Read-only analytics over current bill data. No schema change, no migration, no
DB writes. "Open AP" = open_balance_cents > 0 AND is_paid = 0; sums
open_balance_cents. Same convention used in bills.py / followup.py.

One landscape-printable page, four sections:
  1. Header band  -- Total Open AP, bill count, Uncategorized count, "as of"
  2. Aging        -- Current / 1-30 / 31-60 / 61-90 / 90+ / No due date
  3. Categories   -- by bill_metadata.app_category; Uncategorized pinned bottom
  4. Top vendors  -- top 20 by Open $ + "All other vendors (N)" reconciling row

All three pivots tie to the same grand total (the header band Total Open AP).

Pure helpers (compute_*, build_summary_workbook) take fetched row dicts and
return shape-ready structures, so math + workbook are unit-testable without
Flask or a DB -- the excel_payrun.py pattern.

Excel export: single in-memory .xlsx with 4 sheets (Summary/Aging/Categories/
Top Vendors). Snapshot-on-demand, not written to exports/ and not audited --
this is read-only analytics, not a financial artifact of record like the
Phase 5 CFO/CEO pay-run exports.
"""
from datetime import date, datetime, timezone
from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from flask import Blueprint, render_template, send_file
from flask_login import login_required

import db
import dates as dates_mod
from auth import role_required

bp = Blueprint("summary", __name__)


def init_summary(app):
    app.register_blueprint(bp)


# -- constants --------------------------------------------------------------

UNCATEGORIZED = "Uncategorized"
TOP_VENDORS_N = 20

# Display order. NO_DUE_DATE_LABEL is a separate row appended below the buckets.
AGING_LABELS = ["Current", "1–30", "31–60", "61–90", "90+"]
NO_DUE_DATE_LABEL = "No due date"
AGING_ROWS = AGING_LABELS + [NO_DUE_DATE_LABEL]

# Phase 4.6: Pipeline bucketing by expected_payment_date (the CFO cash-forecast
# field). Rolling day-windows from today; a PAST expected date folds into "This
# week" (it's an imminent/overdue cash need). NO_EXPECTED_LABEL is appended last.
PIPELINE_LABELS = ["This week", "This month", "Next month", "Later"]
NO_EXPECTED_LABEL = "No date set"
PIPELINE_ROWS = PIPELINE_LABELS + [NO_EXPECTED_LABEL]

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# -- DB queries -------------------------------------------------------------

def _fetch_open_bills(conn):
    """Open AP rows. LEFT JOIN bill_metadata so a missing metadata row degrades
    to NULL app_category (folded into Uncategorized) instead of dropping the bill.
    Phase 4.6: also pulls the 2-D classification + the two date fields; a missing
    metadata row COALESCEs to the column defaults (ordinary_ap / not_due) so every
    open bill lands in exactly one section (no bill falls through the cracks)."""
    rows = conn.execute(
        "SELECT b.qb_bill_id, b.vendor, b.bill_date, b.due_date, "
        "       b.open_balance_cents, m.app_category, "
        "       COALESCE(m.obligation_type,'ordinary_ap') AS obligation_type, "
        "       COALESCE(m.due_state,'not_due')           AS due_state, "
        "       m.invoice_due_date, m.expected_payment_date "
        "FROM bill b LEFT JOIN bill_metadata m ON m.qb_bill_id = b.qb_bill_id "
        "WHERE b.open_balance_cents > 0 AND b.is_paid = 0"
    ).fetchall()
    return [dict(r) for r in rows]


def _last_sync_at(conn):
    """Canonical 'as of' timestamp: latest audit_log action='sync_run' row's
    created_at (the same source admin._latest_sync uses). Falls back to
    MAX(bill.last_synced_at). Returns ISO string or None."""
    r = conn.execute(
        "SELECT created_at FROM audit_log WHERE action='sync_run' "
        "ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if r and r["created_at"]:
        return r["created_at"]
    r = conn.execute("SELECT MAX(last_synced_at) AS m FROM bill").fetchone()
    return r["m"] if r and r["m"] else None


# -- pure pivot computations ------------------------------------------------

def _open_total(rows):
    return sum((r.get("open_balance_cents") or 0) for r in rows)


def aging_bucket(due_iso, today):
    """Bucket label for a due_date. NULL -> NO_DUE_DATE_LABEL; dpd<=0 -> Current;
    1..30, 31..60, 61..90 -> their bands; dpd>=91 -> '90+'."""
    if not due_iso:
        return NO_DUE_DATE_LABEL
    dd = dates_mod._as_date(due_iso)
    if dd is None:
        return NO_DUE_DATE_LABEL
    dpd = (today - dd).days
    if dpd <= 0:
        return "Current"
    if dpd <= 30:
        return "1–30"
    if dpd <= 60:
        return "31–60"
    if dpd <= 90:
        return "61–90"
    return "90+"


def compute_header(rows, last_sync_iso):
    return {
        "total_open_cents": _open_total(rows),
        "bill_count": len(rows),
        "uncategorized_count": sum(
            1 for r in rows
            if (r.get("app_category") or UNCATEGORIZED) == UNCATEGORIZED
        ),
        "as_of": last_sync_iso,
    }


def compute_aging_buckets(rows, today, date_key="due_date"):
    """Returns 6 rows in display order: [{label, bill_count, open_cents, pct}].
    Sums to the grand total (modulo rounding on pct). `date_key` selects the
    date field to age on -- 'due_date' by default; Right Now AP ages on
    'invoice_due_date' (the contractual, QB-locked date)."""
    total = _open_total(rows)
    agg = {lbl: {"bill_count": 0, "open_cents": 0} for lbl in AGING_ROWS}
    for r in rows:
        lbl = aging_bucket(r.get(date_key), today)
        agg[lbl]["bill_count"] += 1
        agg[lbl]["open_cents"] += r.get("open_balance_cents") or 0
    out = []
    for lbl in AGING_ROWS:
        a = agg[lbl]
        pct = (a["open_cents"] / total * 100.0) if total else 0.0
        out.append({"label": lbl, "bill_count": a["bill_count"],
                    "open_cents": a["open_cents"], "pct": pct})
    return out


def compute_category_concentration(rows):
    """Group by app_category (NULL -> 'Uncategorized'). Sort Open $ desc;
    Uncategorized pinned to the bottom so the eye lands on real categories
    but it stays visible as a hygiene flag."""
    total = _open_total(rows)
    agg = {}
    for r in rows:
        cat = r.get("app_category") or UNCATEGORIZED
        a = agg.setdefault(cat, {"bill_count": 0, "open_cents": 0})
        a["bill_count"] += 1
        a["open_cents"] += r.get("open_balance_cents") or 0
    items = sorted(agg.items(), key=lambda kv: kv[1]["open_cents"], reverse=True)
    items = ([kv for kv in items if kv[0] != UNCATEGORIZED]
             + [kv for kv in items if kv[0] == UNCATEGORIZED])
    out = []
    for cat, a in items:
        pct = (a["open_cents"] / total * 100.0) if total else 0.0
        out.append({"label": cat, "bill_count": a["bill_count"],
                    "open_cents": a["open_cents"], "pct": pct})
    return out


def compute_top_vendors(rows, n=TOP_VENDORS_N):
    """Top N vendors by Open $ desc, each with oldest_bill_date. Appends an
    'All other vendors (N)' reconciling row so the column ties to the grand
    total. Returns [{label, bill_count, oldest_bill_date, open_cents, pct,
    vendor_exact, other_vendor_count}]; vendor_exact is the literal bill.vendor
    string for drill-down (None on the reconciling row)."""
    total = _open_total(rows)
    agg = {}
    for r in rows:
        v = r.get("vendor") or ""
        a = agg.setdefault(v, {"bill_count": 0, "open_cents": 0, "oldest": None})
        a["bill_count"] += 1
        a["open_cents"] += r.get("open_balance_cents") or 0
        bd = dates_mod._as_date(r.get("bill_date"))
        if bd is not None and (a["oldest"] is None or bd < a["oldest"]):
            a["oldest"] = bd
    items = sorted(agg.items(), key=lambda kv: kv[1]["open_cents"], reverse=True)
    top, rest = items[:n], items[n:]
    out = []
    for v, a in top:
        pct = (a["open_cents"] / total * 100.0) if total else 0.0
        out.append({
            "label": v or "(blank)",
            "bill_count": a["bill_count"],
            "oldest_bill_date": a["oldest"].isoformat() if a["oldest"] else None,
            "open_cents": a["open_cents"], "pct": pct,
            "vendor_exact": v or None,
            "other_vendor_count": None,
        })
    if rest:
        bc = sum(a["bill_count"] for _, a in rest)
        oc = sum(a["open_cents"] for _, a in rest)
        pct = (oc / total * 100.0) if total else 0.0
        out.append({
            "label": f"All other vendors ({len(rest)})",
            "bill_count": bc, "oldest_bill_date": None,
            "open_cents": oc, "pct": pct,
            "vendor_exact": None,
            "other_vendor_count": len(rest),
        })
    return out


# -- Phase 4.6: 2-D classification sections ---------------------------------

def is_right_now_ap(r):
    """Board-defensible AP: a real vendor invoice that is payable right now."""
    return r.get("obligation_type") == "ordinary_ap" and r.get("due_state") == "due"


def is_pipeline(r):
    """Coming / contingent: real AP waiting on a trigger, plus not-real-AP
    (deposits, placeholders). Excludes debt service."""
    obl, due = r.get("obligation_type"), r.get("due_state")
    return (obl == "ordinary_ap" and due == "not_due") or obl == "not_real_ap"


def is_debt_service(r):
    """Liability paydown. Pay-run eligible when due, but NEVER folded into AP."""
    return r.get("obligation_type") == "debt_service"


def split_sections(rows):
    """Partition open-AP rows into the three dashboard sections. Every row lands
    in exactly one (the cells are mutually exclusive and exhaustive over the
    valid 2-D combinations; not_real_ap is forced not_due upstream)."""
    return {
        "right_now": [r for r in rows if is_right_now_ap(r)],
        "pipeline": [r for r in rows if is_pipeline(r)],
        "debt_service": [r for r in rows if is_debt_service(r)],
    }


def pipeline_bucket(expected_iso, today):
    """Bucket an expected_payment_date. NULL -> 'No date set'. A past/overdue date
    folds into 'This week'. Windows: <=7d week, <=30d month, <=60d next month,
    else later."""
    if not expected_iso:
        return NO_EXPECTED_LABEL
    d = dates_mod._as_date(expected_iso)
    if d is None:
        return NO_EXPECTED_LABEL
    delta = (d - today).days
    if delta <= 7:
        return "This week"
    if delta <= 30:
        return "This month"
    if delta <= 60:
        return "Next month"
    return "Later"


def compute_right_now_ap(rows, today):
    """Right Now AP section: ordinary_ap + due, aged by invoice_due_date.
    Returns {total_cents, bill_count, aging:[6 rows], by_category:[...]}."""
    rnap = [r for r in rows if is_right_now_ap(r)]
    return {
        "total_cents": _open_total(rnap),
        "bill_count": len(rnap),
        "aging": compute_aging_buckets(rnap, today, date_key="invoice_due_date"),
        "by_category": compute_category_concentration(rnap),
    }


def compute_pipeline(rows, today):
    """Pipeline / Contingent section: ordinary_ap+not_due plus not_real_ap,
    bucketed by expected_payment_date. Returns {total_cents, bill_count,
    buckets:[5 rows], waiting_cents, waiting_n, not_real_cents, not_real_n}.
    The two sub-totals split the 'waiting on a trigger' (real AP, not yet due)
    from the 'not real AP' (deposits/placeholders) money inside the section."""
    pipe = [r for r in rows if is_pipeline(r)]
    total = _open_total(pipe)
    agg = {lbl: {"bill_count": 0, "open_cents": 0} for lbl in PIPELINE_ROWS}
    for r in pipe:
        lbl = pipeline_bucket(r.get("expected_payment_date"), today)
        agg[lbl]["bill_count"] += 1
        agg[lbl]["open_cents"] += r.get("open_balance_cents") or 0
    buckets = []
    for lbl in PIPELINE_ROWS:
        a = agg[lbl]
        pct = (a["open_cents"] / total * 100.0) if total else 0.0
        buckets.append({"label": lbl, "bill_count": a["bill_count"],
                        "open_cents": a["open_cents"], "pct": pct})
    waiting = [r for r in pipe if r.get("obligation_type") == "ordinary_ap"]
    not_real = [r for r in pipe if r.get("obligation_type") == "not_real_ap"]
    return {
        "total_cents": total, "bill_count": len(pipe), "buckets": buckets,
        "waiting_cents": _open_total(waiting), "waiting_n": len(waiting),
        "not_real_cents": _open_total(not_real), "not_real_n": len(not_real),
    }


def compute_debt_service(rows):
    """Debt Service section: every debt_service bill as an upcoming installment,
    soonest invoice_due_date first (NULLs last). Split due vs not_due. NOT part
    of any AP total. Returns {total_cents, due_cents, not_due_cents, bill_count,
    installments:[{vendor, invoice_due_date, due_state, open_cents}]}."""
    debt = [r for r in rows if is_debt_service(r)]
    due = [r for r in debt if r.get("due_state") == "due"]
    installments = sorted(
        debt, key=lambda r: (r.get("invoice_due_date") is None,
                             r.get("invoice_due_date") or ""))
    return {
        "total_cents": _open_total(debt),
        "due_cents": _open_total(due),
        "not_due_cents": _open_total([r for r in debt if r.get("due_state") != "due"]),
        "bill_count": len(debt),
        "installments": [{
            "vendor": r.get("vendor") or "(blank)",
            "invoice_due_date": r.get("invoice_due_date"),
            "due_state": r.get("due_state"),
            "open_cents": r.get("open_balance_cents") or 0,
        } for r in installments],
    }


# -- workbook builder (pure) ------------------------------------------------

_MONEY_FMT = "#,##0.00"
_PCT_FMT = "0.0%"
_DATE_FMT = "mm/dd/yyyy"

_HDR_FONT = Font(name="Arial", size=10, bold=True)
_BODY_FONT = Font(name="Arial", size=10)
_TOTAL_FONT = Font(name="Arial", size=11, bold=True)
_TITLE_FONT = Font(name="Arial", size=12, bold=True)
_FOOT_FONT = Font(name="Arial", size=8, italic=True)


def _setup_landscape(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"
    ws.oddFooter.center.text = "Page &P of &N"


def _now_iso():
    return datetime.now(timezone.utc).isoformat(sep=" ", timespec="seconds")


def _write_table(ws, headers, rows, widths, *, money_cols=(), pct_cols=(),
                 date_cols=(), total_row=None):
    """Generic 1-indexed table writer starting at row 1. money/pct/date cols
    are 1-based indices. total_row is an optional bold footer."""
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = _HDR_FONT
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 2
    for row in rows:
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.font = _BODY_FONT
            if c in money_cols:
                cell.number_format = _MONEY_FMT
            elif c in pct_cols:
                cell.number_format = _PCT_FMT
            elif c in date_cols and v is not None:
                cell.number_format = _DATE_FMT
        r += 1
    if total_row is not None:
        for c, v in enumerate(total_row, 1):
            cell = ws.cell(r, c, v); cell.font = _TOTAL_FONT
            if c in money_cols:
                cell.number_format = _MONEY_FMT
            elif c in pct_cols:
                cell.number_format = _PCT_FMT
        r += 1
    ws.freeze_panes = "A2"


def build_summary_workbook(data):
    """data keys: as_of, right_now, pipeline, debt_service, today_iso.
    Returns an openpyxl.Workbook with 5 sheets: Overview, Right Now AP, RNAP by
    Category, Pipeline, Debt Service. All sheets set up for landscape print.
    Mirrors the /summary three-section dashboard."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    rn, pipe, debt = data["right_now"], data["pipeline"], data["debt_service"]

    # ----- Overview (the three headline totals) -----
    ws = wb.create_sheet("Overview")
    _setup_landscape(ws)
    ws.cell(1, 1, "AP Summary — by classification").font = _TITLE_FONT
    ws.cell(2, 1, f"As of: {data.get('as_of') or 'never synced'}").font = _BODY_FONT
    ws.cell(3, 1, f"Generated: {_now_iso()}").font = _FOOT_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    rowspec = [
        ("Right Now AP (ordinary, due)", rn["total_cents"]),
        ("Pipeline / Contingent AP", pipe["total_cents"]),
        ("Debt Service (NOT in AP)", debt["total_cents"]),
    ]
    r = 5
    for label, cents in rowspec:
        ws.cell(r, 1, label).font = _HDR_FONT
        v = ws.cell(r, 2, cents / 100.0); v.font = _TOTAL_FONT; v.number_format = _MONEY_FMT
        r += 1
    ws.cell(r + 1, 1, "Right Now AP excludes debt service (a liability paydown) "
            "and not-real-AP (no obligation yet).").font = _FOOT_FONT

    # ----- Right Now AP (aging by invoice_due_date) -----
    ws = wb.create_sheet("Right Now AP")
    _setup_landscape(ws)
    rows = [[b["label"], b["bill_count"], b["open_cents"] / 100.0, b["pct"] / 100.0]
            for b in rn["aging"]]
    tot_c = sum(b["open_cents"] for b in rn["aging"])
    tot_n = sum(b["bill_count"] for b in rn["aging"])
    _write_table(
        ws, ["Aging (by invoice due date)", "Bill count", "Open $", "% of total"], rows,
        widths=[28, 12, 16, 12], money_cols=(3,), pct_cols=(4,),
        total_row=["TOTAL", tot_n, tot_c / 100.0, 1.0 if tot_c else 0.0],
    )

    # ----- RNAP by Category -----
    ws = wb.create_sheet("RNAP by Category")
    _setup_landscape(ws)
    rows = [[c["label"], c["bill_count"], c["open_cents"] / 100.0, c["pct"] / 100.0]
            for c in rn["by_category"]]
    tot_c = sum(c["open_cents"] for c in rn["by_category"])
    tot_n = sum(c["bill_count"] for c in rn["by_category"])
    _write_table(
        ws, ["Category (Right Now AP)", "Bill count", "Open $", "% of total"], rows,
        widths=[40, 12, 16, 12], money_cols=(3,), pct_cols=(4,),
        total_row=["TOTAL", tot_n, tot_c / 100.0, 1.0 if tot_c else 0.0],
    )

    # ----- Pipeline (buckets by expected_payment_date) -----
    ws = wb.create_sheet("Pipeline")
    _setup_landscape(ws)
    rows = [[b["label"], b["bill_count"], b["open_cents"] / 100.0, b["pct"] / 100.0]
            for b in pipe["buckets"]]
    tot_c = sum(b["open_cents"] for b in pipe["buckets"])
    tot_n = sum(b["bill_count"] for b in pipe["buckets"])
    _write_table(
        ws, ["Coming due (by expected pay date)", "Bill count", "Open $", "% of total"],
        rows, widths=[30, 12, 16, 12], money_cols=(3,), pct_cols=(4,),
        total_row=["TOTAL", tot_n, tot_c / 100.0, 1.0 if tot_c else 0.0],
    )

    # ----- Debt Service (installments) -----
    ws = wb.create_sheet("Debt Service")
    _setup_landscape(ws)
    rows = []
    for it in debt["installments"]:
        rows.append([it["vendor"], _xl_date(it["invoice_due_date"]),
                     (it["due_state"] or "").replace("_", " "), it["open_cents"] / 100.0])
    _write_table(
        ws, ["Vendor", "Invoice due date", "Due state", "Open $"], rows,
        widths=[36, 16, 12, 16], money_cols=(4,), date_cols=(2,),
        total_row=["TOTAL (not part of AP)", debt["bill_count"], None,
                   debt["total_cents"] / 100.0],
    )
    return wb


def _xl_date(s):
    """ISO date string/date/datetime -> date for a real Excel date; None if blank."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    try:
        return date.fromisoformat(str(s)[:10])
    except ValueError:
        return None


# -- BOD AP report (Phase 4.6): the board-defensible ordinary_ap aging report --

def build_bod_ap_workbook(rows, today, as_of):
    """Board AP report. ordinary_ap ONLY, in two sections -- Due (Right Now AP,
    aged by invoice_due_date) and Not Due (aged by expected_payment_date) --
    each with its own total, plus a combined ordinary-AP total. EXCLUDES
    debt_service (booked as a liability) and not_real_ap (no obligation yet):
    the asymmetry the board number depends on."""
    ordinary = [r for r in rows if r.get("obligation_type") == "ordinary_ap"]
    due = [r for r in ordinary if r.get("due_state") == "due"]
    not_due = [r for r in ordinary if r.get("due_state") != "due"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("BOD AP Report")
    _setup_landscape(ws)

    def _aging_rows(subset, date_key):
        buckets = compute_aging_buckets(subset, today, date_key=date_key)
        return [[b["label"], b["bill_count"], b["open_cents"] / 100.0] for b in buckets], \
               sum(b["open_cents"] for b in buckets), sum(b["bill_count"] for b in buckets)

    for c, h in enumerate(["AP Aging (ordinary AP only)", "Bill count", "Open $"], 1):
        ws.cell(1, c, h).font = _HDR_FONT
    for c, w in enumerate([34, 12, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.cell(2, 1, f"As of: {as_of or 'never synced'} · Generated {_now_iso()}").font = _FOOT_FONT

    r = 4
    for title, subset, key in [("DUE NOW (aged by invoice due date)", due, "invoice_due_date"),
                               ("NOT YET DUE (aged by expected pay date)", not_due,
                                "expected_payment_date")]:
        ws.cell(r, 1, title).font = _TITLE_FONT
        r += 1
        body, sub_c, sub_n = _aging_rows(subset, key)
        for label, n, c_ in body:
            ws.cell(r, 1, label).font = _BODY_FONT
            ws.cell(r, 2, n).font = _BODY_FONT
            mc = ws.cell(r, 3, c_); mc.font = _BODY_FONT; mc.number_format = _MONEY_FMT
            r += 1
        ws.cell(r, 1, "Section total").font = _TOTAL_FONT
        ws.cell(r, 2, sub_n).font = _TOTAL_FONT
        tc = ws.cell(r, 3, sub_c / 100.0); tc.font = _TOTAL_FONT; tc.number_format = _MONEY_FMT
        r += 2

    ws.cell(r, 1, "ORDINARY AP TOTAL (due + not due)").font = _TOTAL_FONT
    ws.cell(r, 2, len(ordinary)).font = _TOTAL_FONT
    gc = ws.cell(r, 3, _open_total(ordinary) / 100.0)
    gc.font = _TOTAL_FONT; gc.number_format = _MONEY_FMT
    r += 2
    ws.cell(r, 1, "Excludes debt service and not-real-AP by design.").font = _FOOT_FONT
    return wb


# -- routes -----------------------------------------------------------------

def _gather(conn, today):
    """The three-section dashboard data (Phase 4.6). Every open-AP bill lands in
    exactly one of Right Now AP / Pipeline / Debt Service."""
    rows = _fetch_open_bills(conn)
    return {
        "as_of": _last_sync_at(conn),
        "right_now": compute_right_now_ap(rows, today),
        "pipeline": compute_pipeline(rows, today),
        "debt_service": compute_debt_service(rows),
        "today_iso": today.isoformat(),
    }


@bp.route("/summary")
@login_required
def summary():
    data = _gather(db.get_db(), date.today())
    return render_template("summary.html", **data)


@bp.route("/summary/export.xlsx")
@login_required
def export_xlsx():
    today = date.today()
    data = _gather(db.get_db(), today)
    wb = build_summary_workbook(data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    filename = f"MRP_AP_Summary_{today.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename,
                     mimetype=_XLSX_MIME)


@bp.route("/summary/bod-ap-report.xlsx")
@role_required("controller", "cfo")
def export_bod_ap():
    """The board AP report: ordinary_ap aging (due + not-due sections), excludes
    debt_service and not_real_ap. Controller + CFO only -- this is the artifact
    the CFO defends to the board."""
    today = date.today()
    conn = db.get_db()
    rows = _fetch_open_bills(conn)
    wb = build_bod_ap_workbook(rows, today, _last_sync_at(conn))
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    filename = f"MRP_BOD_AP_Report_{today.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename,
                     mimetype=_XLSX_MIME)
