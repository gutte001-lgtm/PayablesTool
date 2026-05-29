"""
test_phase_6_summary.py -- Phase 6 spend summary tests (/summary + export).

Plain-script style (mirrors test_phase_5_export.py): check(label, cond); exit
code == failures; run with `python test_phase_6_summary.py`.

PART A (pure): aging boundaries, NULL due-date routing, category sort with
Uncategorized pinned bottom, top-20 + reconciling "all other vendors" row,
all three pivots tie to one grand total, build_summary_workbook produces a
4-sheet workbook with totals on each detail sheet.

PART B (routes): temp DB + Flask test client -- /summary renders, paid bills
are excluded from Open AP, the export route returns a valid 4-sheet xlsx with
the MRP_AP_Summary_YYYY-MM-DD filename, anonymous users are redirected to
login, and the two bills.py filters the summary drill-down links depend on
(?vendor= exact, ?uncat=1) actually filter the bill list. Skipped if no
SECRET_KEY in .env. Live payables.db is never touched.
"""
import io
import re
import shutil
import sqlite3
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

FAILURES = []
_TMP = []


def check(label, cond):
    print(("ok   " if cond else "FAIL ") + label)
    if not cond:
        FAILURES.append(label)


ROOT = Path(__file__).resolve().parent
import summary       # noqa: E402
import openpyxl      # noqa: E402


def row(bid, vendor, open_cents, due_offset=None, bill_offset=-30, cat="Parts & Products"):
    """Build a bill row dict. due_offset: days from `today` (None -> NULL due_date).
    bill_offset: days from today for bill_date (default 30 days ago)."""
    today = date(2026, 5, 26)
    due = (today + timedelta(days=due_offset)).isoformat() if due_offset is not None else None
    return {
        "qb_bill_id": bid, "vendor": vendor, "open_balance_cents": open_cents,
        "due_date": due,
        "bill_date": (today + timedelta(days=bill_offset)).isoformat(),
        "app_category": cat,
    }


TODAY = date(2026, 5, 26)

# ====================================================================
print("=" * 60); print("PART A -- pure pivots + workbook"); print("=" * 60)

# ---- aging boundaries ----
check("aging: due today (dpd=0) -> Current",
      summary.aging_bucket(TODAY.isoformat(), TODAY) == "Current")
check("aging: due tomorrow (dpd=-1) -> Current",
      summary.aging_bucket((TODAY + timedelta(days=1)).isoformat(), TODAY) == "Current")
check("aging: 1d past due -> 1-30",
      summary.aging_bucket((TODAY - timedelta(days=1)).isoformat(), TODAY) == "1–30")
check("aging: 30d past due -> 1-30",
      summary.aging_bucket((TODAY - timedelta(days=30)).isoformat(), TODAY) == "1–30")
check("aging: 31d past due -> 31-60",
      summary.aging_bucket((TODAY - timedelta(days=31)).isoformat(), TODAY) == "31–60")
check("aging: 60d past due -> 31-60",
      summary.aging_bucket((TODAY - timedelta(days=60)).isoformat(), TODAY) == "31–60")
check("aging: 61d past due -> 61-90",
      summary.aging_bucket((TODAY - timedelta(days=61)).isoformat(), TODAY) == "61–90")
check("aging: 90d past due -> 61-90",
      summary.aging_bucket((TODAY - timedelta(days=90)).isoformat(), TODAY) == "61–90")
check("aging: 91d past due -> 90+",
      summary.aging_bucket((TODAY - timedelta(days=91)).isoformat(), TODAY) == "90+")
check("aging: 9999d past due -> 90+",
      summary.aging_bucket((TODAY - timedelta(days=9999)).isoformat(), TODAY) == "90+")
check("aging: NULL due_date -> No due date",
      summary.aging_bucket(None, TODAY) == "No due date")
check("aging: empty string -> No due date",
      summary.aging_bucket("", TODAY) == "No due date")

# ---- aging buckets: distribution + ties to total ----
ROWS = [
    row("b1", "Acme",    100_00, due_offset=5),     # Current
    row("b2", "Acme",    200_00, due_offset=-15),   # 1-30
    row("b3", "Beta",    300_00, due_offset=-45),   # 31-60
    row("b4", "Gamma",   400_00, due_offset=-75),   # 61-90
    row("b5", "Delta",   500_00, due_offset=-200),  # 90+
    row("b6", "Epsilon", 600_00, due_offset=None),  # No due date
    row("b7", "Acme",    700_00, due_offset=10),    # Current
]
GRAND = sum(r["open_balance_cents"] for r in ROWS)  # 2800_00

ag = summary.compute_aging_buckets(ROWS, TODAY)
labels = [b["label"] for b in ag]
check("aging: 6 rows in display order",
      labels == ["Current", "1–30", "31–60", "61–90", "90+", "No due date"])
by_lbl = {b["label"]: b for b in ag}
check("aging: Current = b1+b7 ($800)", by_lbl["Current"]["open_cents"] == 800_00)
check("aging: 1-30 = b2 ($200)", by_lbl["1–30"]["open_cents"] == 200_00)
check("aging: 31-60 = b3 ($300)", by_lbl["31–60"]["open_cents"] == 300_00)
check("aging: 61-90 = b4 ($400)", by_lbl["61–90"]["open_cents"] == 400_00)
check("aging: 90+ = b5 ($500)", by_lbl["90+"]["open_cents"] == 500_00)
check("aging: No due date = b6 ($600)", by_lbl["No due date"]["open_cents"] == 600_00)
check("aging: sum of 6 rows ties to grand total",
      sum(b["open_cents"] for b in ag) == GRAND)
check("aging: percentages sum to ~100",
      abs(sum(b["pct"] for b in ag) - 100.0) < 0.01)

# ---- categories: NULL -> Uncategorized, sort desc, Uncategorized pinned bottom ----
CAT_ROWS = [
    row("c1", "V1", 1000_00, due_offset=0, cat="Notes Payable"),       # biggest
    row("c2", "V2", 500_00,  due_offset=0, cat="New Device Purchases"),
    row("c3", "V3", 300_00,  due_offset=0, cat="Notes Payable"),       # merges with c1
    row("c4", "V4", 9999_00, due_offset=0, cat=None),                  # NULL -> Uncategorized (biggest $)
    row("c5", "V5", 50_00,   due_offset=0, cat="Freight"),
]
cats = summary.compute_category_concentration(CAT_ROWS)
cat_labels = [c["label"] for c in cats]
check("cat: NULL cat folded to 'Uncategorized'",
      "Uncategorized" in cat_labels and None not in cat_labels)
check("cat: Uncategorized is the LAST row (pinned bottom even though biggest)",
      cat_labels[-1] == "Uncategorized")
non_uncat = cat_labels[:-1]
check("cat: non-Uncategorized sorted by Open $ desc",
      non_uncat == ["Notes Payable", "New Device Purchases", "Freight"])
by_cat = {c["label"]: c for c in cats}
check("cat: Notes Payable merges to $1300 (2 bills)",
      by_cat["Notes Payable"]["open_cents"] == 1300_00 and by_cat["Notes Payable"]["bill_count"] == 2)
check("cat: sum across all categories ties to grand total",
      sum(c["open_cents"] for c in cats) == sum(r["open_balance_cents"] for r in CAT_ROWS))

# ---- top vendors: top-20 + reconciling "All other vendors" row ----
# 22 vendors so we get a remainder of 2
VROWS = []
for i in range(22):
    # vendor i has open = (1000 - i*10) cents, so largest is v0 = 1000, smallest = 790
    VROWS.append(row(f"vb{i}", f"Vendor{i:02d}", 1000_00 - i * 1000, due_offset=0,
                     bill_offset=-(i + 1)))   # oldest = the biggest vendor

tv = summary.compute_top_vendors(VROWS, n=20)
check("top: 21 rows = 20 top + 1 'All other vendors' reconciling",
      len(tv) == 21 and tv[-1]["label"] == "All other vendors (2)")
check("top: top vendor is Vendor00 (largest open)",
      tv[0]["label"] == "Vendor00" and tv[0]["vendor_exact"] == "Vendor00")
check("top: reconciling row has vendor_exact=None (no drill-down link)",
      tv[-1]["vendor_exact"] is None and tv[-1]["other_vendor_count"] == 2)
check("top: sum of all 21 rows ties to grand total",
      sum(v["open_cents"] for v in tv) == sum(r["open_balance_cents"] for r in VROWS))
check("top: oldest_bill_date present on top rows, None on reconciling row",
      tv[0]["oldest_bill_date"] is not None and tv[-1]["oldest_bill_date"] is None)

# fewer than 20 vendors -> no reconciling row
tv_few = summary.compute_top_vendors(CAT_ROWS, n=20)
check("top: <20 vendors -> no 'All other vendors' row",
      not any("All other vendors" in v["label"] for v in tv_few))

# ---- all three pivots tie to ONE grand total ----
hdr = summary.compute_header(ROWS, "2026-05-26 12:00:00")
ag2 = summary.compute_aging_buckets(ROWS, TODAY)
cat2 = summary.compute_category_concentration(ROWS)
tv2 = summary.compute_top_vendors(ROWS)
check("ties: header.total == aging total == categories total == vendors total",
      hdr["total_open_cents"] == sum(b["open_cents"] for b in ag2)
      == sum(c["open_cents"] for c in cat2)
      == sum(v["open_cents"] for v in tv2))

# ---- header band: total, count, uncategorized ----
hdr3 = summary.compute_header(CAT_ROWS, "2026-05-26 12:00:00")
check("hdr: bill_count = 5", hdr3["bill_count"] == 5)
check("hdr: uncategorized_count = 1 (the NULL cat row)",
      hdr3["uncategorized_count"] == 1)
check("hdr: total_open_cents sums all", hdr3["total_open_cents"] == 11849_00)

# NOTE: the build_summary_workbook structure changed in Phase 4.6 (the dashboard
# was restructured into the three classification sections). The workbook shape
# (Overview / Right Now AP / RNAP by Category / Pipeline / Debt Service) is now
# covered by test_phase_4_6.py. The generic pivot functions tested above
# (aging/category/vendor/header) are unchanged and still in use.

# ====================================================================
print("\n" + "=" * 60); print("PART B -- routes (temp DB + client)"); print("=" * 60)
from dotenv import dotenv_values  # noqa: E402
if not dotenv_values(ROOT / ".env").get("SECRET_KEY"):
    print("SKIP: no SECRET_KEY in .env (pure tests above still ran)")
else:
    from werkzeug.security import generate_password_hash  # noqa: E402
    import db          # noqa: E402
    import init_db     # noqa: E402
    from app import app  # noqa: E402
    app.config["WTF_CSRF_ENABLED"] = False
    app.config["TESTING"] = True
    PW = generate_password_hash("testpw")
    USERS = [("marilyn", "Marilyn", "ap_clerk"), ("joe", "Joe", "controller"),
             ("shaun", "Shaun", "cfo")]

    def fresh_db():
        d = Path(tempfile.mkdtemp()); _TMP.append(d)
        db.DB_PATH = d / "p6.db"
        cn = sqlite3.connect(db.DB_PATH); cn.executescript(init_db.SCHEMA)
        for u, name, role in USERS:
            cn.execute("INSERT INTO users (username,name,role,password_hash,is_active) "
                       "VALUES (?,?,?,?,1)", (u, name, role, PW))
        cn.commit(); cn.close()
        return d

    def _cn():
        c = sqlite3.connect(db.DB_PATH); c.row_factory = sqlite3.Row; return c

    def login(u):
        c = app.test_client(); c.post("/login", data={"username": u, "password": "testpw"}); return c

    def seed_bill(bid, vendor, open_cents, cat=None, due_offset=0, paid=False,
                  obligation="ordinary_ap", due_state="due"):
        """When paid=True: is_paid=1 and open_balance_cents=0 (the warehouse
        sets balance to 0 when QB marks a bill paid; sync flips is_paid).
        amount_cents stays at the original bill amount. obligation/due_state
        (Phase 4.6) default to ordinary_ap+due so a seeded bill is Right Now AP."""
        c = _cn()
        due = (TODAY + timedelta(days=due_offset)).isoformat() if due_offset is not None else None
        c.execute("INSERT INTO bill (qb_bill_id,vendor,bill_number,amount_cents,"
                  "open_balance_cents,bill_date,due_date,is_paid,last_synced_at) "
                  "VALUES (?,?,?,?,?,?,?,?,?)",
                  (bid, vendor, "B-" + bid, open_cents,
                   0 if paid else open_cents,
                   "2026-04-01", due,
                   1 if paid else 0,
                   "2026-05-22"))
        # Phase 4.6: seed the 2-D classification so the bill lands in a section.
        # Default ordinary_ap + due (Right Now AP), with invoice/expected dates
        # mirroring due_date so the Right Now AP aging (by invoice_due_date)
        # buckets it the same way Phase 6 expected.
        c.execute("INSERT INTO bill_metadata (qb_bill_id,app_category,approval_state,"
                  "obligation_type,due_state,invoice_due_date,expected_payment_date,"
                  "created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
                  (bid, cat, "New", obligation, due_state, due, due,
                   "2026-05-01", "2026-05-01"))
        c.commit(); c.close()

    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ---- summary page renders + carries header band totals ----
    fresh_db()
    seed_bill("s1", "Acme Corp",  10000_00, cat="Notes Payable",   due_offset=-100)  # 90+
    seed_bill("s2", "Beta LLC",   3000_00,  cat="Freight",          due_offset=5)    # Current
    seed_bill("s3", "Gamma Inc",  2000_00,  cat=None,               due_offset=None) # No due date, Uncat
    # paid bill alongside the open three -- Phase 6 must exclude it from Open AP
    seed_bill("p1", "PaidCo Inc", 99999_00, cat="Freight", due_offset=-30, paid=True)
    c = login("joe")
    r = c.get("/summary")
    check("route: /summary 200 for controller", r.status_code == 200)
    body = r.data.decode("utf-8")
    # Phase 4.6: the three classification sections (all 3 open bills default to
    # ordinary_ap+due -> Right Now AP = $15,000).
    check("route: /summary shows the three section headings",
          "Right Now AP" in body and "Pipeline" in body and "Debt Service" in body)
    check("route: /summary Right Now AP total = $15,000",
          "$15,000.00" in body)
    check("route: /summary shows the No due date aging row (s3 has no invoice date)",
          "No due date" in body)
    check("route: /summary shows a TOTAL row in each section table",
          body.count("TOTAL") >= 3)
    # paid-bill exclusion: PaidCo's $99,999 must not appear anywhere.
    check("route: /summary excludes paid bill -- 'PaidCo Inc' absent from page",
          "PaidCo Inc" not in body)
    check("route: /summary excludes paid bill from totals -- no '99,999' in body",
          "99,999" not in body)
    check("route: Right Now AP shows the 3-bill count",
          "3 bills" in body)
    check("route: cfo can see /summary",
          login("shaun").get("/summary").status_code == 200)
    check("route: ap_clerk can see /summary",
          login("marilyn").get("/summary").status_code == 200)
    # anonymous -> redirect to login
    anon = app.test_client()
    ra = anon.get("/summary")
    check("route: anonymous -> 302 redirect to login",
          ra.status_code == 302 and "/login" in (ra.headers.get("Location") or ""))

    # ---- export route returns the valid 5-sheet xlsx with correct filename ----
    re = c.get("/summary/export.xlsx")
    check("route: export 200 + xlsx mimetype",
          re.status_code == 200 and re.headers.get("Content-Type", "").startswith(XLSX))
    cd = re.headers.get("Content-Disposition", "")
    check("route: filename is MRP_AP_Summary_YYYY-MM-DD.xlsx",
          "MRP_AP_Summary_" in cd and cd.endswith(".xlsx") and date.today().isoformat() in cd)
    dwb = openpyxl.load_workbook(io.BytesIO(re.data))
    check("route: downloaded workbook has the 5 Phase 4.6 sheets",
          dwb.sheetnames == ["Overview", "Right Now AP", "RNAP by Category",
                             "Pipeline", "Debt Service"])
    # Right Now AP sheet aging TOTAL = $15,000 (all 3 open bills are ordinary+due)
    rn_ws = dwb["Right Now AP"]
    tot = [rn_ws.cell(r, 3).value for r in range(2, rn_ws.max_row + 1)
           if rn_ws.cell(r, 1).value == "TOTAL"]
    check("route: downloaded Right Now AP TOTAL Open $ = 15000.00", tot == [15000.0])

    # ---- bills.py ?vendor= exact filter (the small cross-file change) ----
    # seed a second bill with same vendor and one with a different vendor
    seed_bill("v1", "Acme Corp", 500_00, cat="Freight", due_offset=10)
    seed_bill("v2", "Other Co",  500_00, cat="Freight", due_offset=10)
    rv = c.get("/bills?vendor=Acme%20Corp")
    rv_body = rv.data.decode("utf-8")
    check("route: bills.py ?vendor=Acme Corp returns 200",
          rv.status_code == 200)
    check("route: ?vendor=Acme Corp matches Acme bills, not 'Other Co'",
          "Acme Corp" in rv_body and "Other Co" not in rv_body)

    # ---- bills.py ?uncat=1 filter (the link the /summary header uses) ----
    # Seed a bill whose metadata.app_category is the literal string 'Uncategorized'
    # (NULL app_category does NOT match m.app_category='Uncategorized' in SQL).
    seed_bill("uc1", "UncatVendor", 700_00, cat="Uncategorized", due_offset=0)
    ru = c.get("/bills?uncat=1")
    ru_body = ru.data.decode("utf-8")
    check("route: bills.py ?uncat=1 returns 200", ru.status_code == 200)
    check("route: ?uncat=1 shows the Uncategorized bill ('UncatVendor')",
          "UncatVendor" in ru_body)
    check("route: ?uncat=1 excludes bills with other categories",
          "Acme Corp" not in ru_body and "Beta LLC" not in ru_body
          and "Other Co" not in ru_body)

# ====================================================================
for d in _TMP:
    shutil.rmtree(d, ignore_errors=True)
print()
if FAILURES:
    print(f"{len(FAILURES)} FAILURE(S): " + "; ".join(FAILURES))
else:
    print("ALL PHASE 6 SUMMARY CHECKS PASSED")
sys.exit(len(FAILURES))
